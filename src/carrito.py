import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
from ttkwidgets.autocomplete import AutocompleteCombobox
from reportlab.lib.pagesizes import letter
from sqlalchemy.orm import sessionmaker
from database import engine, Productos, Categorias, Clientes, Presupuestos, DetallesPresupuestos, Remitos, DetallesRemitos
from openpyxl import load_workbook
import datetime
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
import os
from PIL import Image, ImageTk
from clientes import ClientesApp
from utils.carrito import agregar_al_carrito, actualizar_carrito, agregar_fuera_lista, eliminar_del_carrito
from utils.precios import modificar_precios, deshacer_ultimo_aumento
from utils.productos import update_productos, buscar_producto
from utils.guardar_remitos import guardar_remito

# Sesión de SQLAlchemy para interactuar con la base de datos
Session = sessionmaker(bind=engine)
session = Session()

# Clase para la interfaz
class RemitosApp(tk.Tk):
    def __init__(self):
        # Inicializar la ventana principal
        super().__init__()
        # Título de la ventana
        self.title("Constructora Jose Luis Lopez")
        # Geometría de la ventana
        self.geometry("1300x700")
        # Crear los widgets con la función create_widgets
        self.create_widgets()
        # Lista para almacenar los productos del carrito
        self.carrito = []
        # Lista para almacenar los precios anteriores antes de un aumento
        self.precios_anteriores = []

    def create_widgets(self):
        # Menú lateral
        # Frame para el menú lateral con un ancho de 200 y color de fondo gris
        self.menu_lateral = tk.Frame(self, width=200, bg='gray')
        # Empaquetar el frame en la ventana principal a la izquierda y que se expanda en el eje y
        self.menu_lateral.pack(side='left', fill='y')

        # Botón para mostrar presupuestos
        # ttk.Button es un botón con un estilo mejorado
        # El botón llama a la función mostrar_presupuestos cuando se hace click 
        self.presupuestos_button = ttk.Button(self.menu_lateral, text="Carrito", command=self.mostrar_presupuestos)
        # Empaquetar el botón en el menú lateral
        self.presupuestos_button.pack(padx=10, pady=10)

        # Botón para mostrar clientes
        self.clientes_button = ttk.Button(self.menu_lateral, text="Clientes", command=self.mostrar_clientes)
        self.clientes_button.pack(padx=10, pady=10)

        # Botón para mostrar listas de precios
        self.listas_precios_button = ttk.Button(self.menu_lateral, text="Precios") #, command=self.mostrar_listas_precios)
        self.listas_precios_button.pack(padx=10, pady=10)

        # Frame principal
        self.main_frame = tk.Frame(self)
        # Empaquetar el frame principal en la ventana principal
        self.main_frame.pack(side='right', fill='both', expand=True)

    def mostrar_presupuestos(self):
        # Limpiar el main_frame antes de agregar nuevos widgets para evitar superposiciones
        for widget in self.main_frame.winfo_children():
            # Destruir el widget
            widget.destroy()

        # Etiqueta para el nombre del cliente
        self.cliente_label = ttk.Label(self.main_frame, text="Nombre del Cliente:")
        self.cliente_label.place(x=10, y=10)
        
        # Combobox para seleccionar el cliente
        self.cliente_var = tk.StringVar()
        self.cliente_combobox = AutocompleteCombobox(self.main_frame, textvariable=self.cliente_var)
        # Establecer la lista de opciones para el combobox con los nombres de los clientes
        self.cliente_combobox.set_completion_list(self.obtener_nombres_clientes())
        self.cliente_combobox.place(x=135, y=10)

        # Menú de selección de tablas (listas de precios)
        self.tabla_label = ttk.Label(self.main_frame, text="Seleccionar una lista:")
        # Empaquetar la etiqueta en el main_frame con place
        self.tabla_label.place(x=10, y=70)

        # Combobox para seleccionar la categoría

        # Obtener todas las categorías de la base de datos
        tablas = session.query(Categorias).all()

        tablas = [tabla.nombre for tabla in tablas]

        # StringVar para almacenar la tabla seleccionada
        self.tabla_var = tk.StringVar()
        # Combobox para seleccionar la tabla, textvariable es la variable que almacena la tabla seleccionada
        self.tabla_combobox = AutocompleteCombobox(self.main_frame, textvariable=self.tabla_var)
        # Establecer la lista de opciones para el combobox
        self.tabla_combobox.set_completion_list(tablas)
        # Llamar a la función update_productos cuando se selecciona una tabla
        self.tabla_combobox.bind("<<ComboboxSelected>>", self.update_productos)
        # Colocar el combobox en el main_frame
        self.tabla_combobox.place(x=135, y=70)

        # Botón para aumentar precios
        # El botón llama a la función aumentar_precios cuando se hace click
        self.aumentar_precios_button = ttk.Button(self.main_frame, text="Modificar Precios", command=self.modificar_precios)
        self.aumentar_precios_button.place(x=300, y=68)

        # Botón para deshacer el último aumento
        # El botón llama a la función deshacer_ultimo_aumento cuando se hace click
        self.deshacer_aumento_button = ttk.Button(self.main_frame, text="Deshacer", command=self.deshacer_ultimo_aumento)
        self.deshacer_aumento_button.place(x=420, y=68)

        # Etiqueta de búsqueda
        self.busqueda_label = ttk.Label(self.main_frame, text="Buscar un producto:")
        self.busqueda_label.place(x=10, y=130)

        # Campo de búsqueda
        # StringVar para almacenar el valor ingresado en el campo de búsqueda 
        self.busqueda_var = tk.StringVar()
        # Entry para ingresar el término de búsqueda, textvariable es la variable que almacena el valor ingresado
        self.busqueda_entry = ttk.Entry(self.main_frame, textvariable=self.busqueda_var)
        self.busqueda_entry.place(x=135, y=130)

        # Botón para buscar productos
        # Cargar la imagen de búsqueda y redimensionarla
        original_search_image = Image.open("./icons/search.png")
        resized_search_image = original_search_image.resize((15, 15), Image.LANCZOS)
        self.search_image = ImageTk.PhotoImage(resized_search_image)

        # El botón llama a la función buscar_producto cuando se hace click
        self.buscar_button = ttk.Button(self.main_frame, image=self.search_image, command=self.buscar_producto)
        self.buscar_button.place(x=270, y=128)

        # Tabla de productos
        # Treeview con las columnas Producto y Precio
        self.productos_tree = ttk.Treeview(self.main_frame, columns=('Codigo', 'Linea', 'Producto', 'Precio'), show='headings')
        # Encabezados de las columnas
        self.productos_tree.heading('Codigo', text='Codigo')
        self.productos_tree.heading('Linea', text='Linea')
        self.productos_tree.heading('Producto', text='Producto')
        self.productos_tree.heading('Precio', text='Precio')
        # Ancho de las columnas
        self.productos_tree.column('Codigo', width=50)
        self.productos_tree.column('Linea', width=70)
        self.productos_tree.column('Producto', width=300)
        self.productos_tree.column('Precio', width=100)
        self.productos_tree.place(x=10, y=160)

        # Scrollbar para la tabla de productos
        # Scrollbar en el eje vertical que se conecta con la tabla de productos y se mueve con ella
        scrollbar = ttk.Scrollbar(self.main_frame, orient=tk.VERTICAL, command=self.productos_tree.yview)
        # Configurar la scrollbar para que se mueva junto con la tabla de productos en el eje y (vertical)
        self.productos_tree.configure(yscroll=scrollbar.set)
        scrollbar.place(x=532, y=160, relheight=0.31)

        # Etiqueta de cantidad
        self.cantidad_label = ttk.Label(self.main_frame, text="Cantidad:")
        self.cantidad_label.place(x=10, y=400)

        # Campo de cantidad
        # IntVar para almacenar el valor ingresado en el campo de cantidad 
        self.cantidad_var = tk.IntVar()
        # Entry para ingresar la cantidad, textvariable es la variable que almacena el valor ingresado
        self.cantidad_entry = ttk.Entry(self.main_frame, textvariable=self.cantidad_var)
        self.cantidad_entry.place(x=115, y=400)

        # Etiqueta de descuento
        self.descuento_label = ttk.Label(self.main_frame, text="Descuento (%):")
        self.descuento_label.place(x=10, y=430)

        # Campo de descuento
        # DoubleVar para almacenar el valor ingresado en el campo de descuento
        self.descuento_var = tk.DoubleVar()
        # Entry para ingresar el descuento, textvariable es la variable que almacena el valor ingresado
        self.descuento_entry = ttk.Entry(self.main_frame, textvariable=self.descuento_var)
        self.descuento_entry.place(x=115, y=430)

        # Botón para agregar al carrito
        # El botón llama a la función agregar_al_carrito cuando se hace click
        self.add_button = ttk.Button(self.main_frame, text="Agregar al Carrito", command=self.agregar_al_carrito)
        self.add_button.place(x=260, y=398)

        # Carrito

        # Agregar productos fuera de lista
        # Etiqueta de productos fuera de lista
        self.productos_fuera_lista_label = ttk.Label(self.main_frame, text="PRODUCTOS FUERA DE LISTA")
        self.productos_fuera_lista_label.place(x=765, y=10)

        # Etiqueta para nombre del producto
        self.producto_label = ttk.Label(self.main_frame, text="Producto:")
        self.producto_label.place(x=750, y=40)

        # Campo para ingresar el nombre del producto
        self.producto_var = tk.StringVar()
        self.producto_entry = ttk.Entry(self.main_frame, textvariable=self.producto_var)
        self.producto_entry.place(x=815, y=40)

        # Etiqueta para cantidad del producto
        self.cantidad_fuera_lista_label = ttk.Label(self.main_frame, text="Cantidad:")
        self.cantidad_fuera_lista_label.place(x=750, y=70)

        # Campo para ingresar la cantidad del producto
        self.cantidad_fuera_lista_var = tk.IntVar()
        self.cantidad_fuera_lista_entry = ttk.Entry(self.main_frame, textvariable=self.cantidad_fuera_lista_var)
        self.cantidad_fuera_lista_entry.place(x=815, y=70)

        # Etiqueta para precio del producto
        self.precio_label = ttk.Label(self.main_frame, text="Precio:")
        self.precio_label.place(x=750, y=100)

        # Campo para ingresar el precio del producto
        self.precio_var = tk.DoubleVar()
        self.precio_entry = ttk.Entry(self.main_frame, textvariable=self.precio_var)
        self.precio_entry.place(x=815, y=100)

        # Botón para agregar productos fuera de lista
        self.add_fuera_lista_button = ttk.Button(self.main_frame, text="Agregar al Carrito", command=self.agregar_fuera_lista)
        self.add_fuera_lista_button.place(x=960, y=68)

        # Botón para borrar del carrito
        # El botón llama a la función eliminar_del_carrito cuando se hace click
        self.delete_button = ttk.Button(self.main_frame, text="Eliminar Producto", command=self.eliminar_del_carrito)
        self.delete_button.place(x=625, y=130)

        # Etiqueta "CARRITO"
        self.carrito_label = ttk.Label(self.main_frame, text="CARRITO")
        self.carrito_label.place(x=810, y=133)

        # Botón para limpiar el carrito
        self.clear_button = ttk.Button(self.main_frame, text="Limpiar Carrito", command=self.limpiar_carrito)
        self.clear_button.place(x=950, y=130)

        # Treeview para mostrar los productos del carrito
        self.carrito_treeview = ttk.Treeview(self.main_frame, columns=('Producto', 'Cantidad', 'Descuento', 'Precio'), show='headings')
        # Encabezados de las columnas
        self.carrito_treeview.heading('Producto', text='Producto')
        self.carrito_treeview.heading('Cantidad', text='Cantidad')
        self.carrito_treeview.heading('Descuento', text='Descuento')
        self.carrito_treeview.heading('Precio', text='Precio')
        # Ancho de las columnas
        self.carrito_treeview.column('Producto', width=350)
        self.carrito_treeview.column('Cantidad', width=75)
        self.carrito_treeview.column('Descuento', width=75)
        self.carrito_treeview.column('Precio', width=100)
        self.carrito_treeview.place(x=552, y=160)

        self.actualizar_carrito()

        # Scrollbar para la lista de productos del carrito
        # Scrollbar en el eje vertical que se conecta con la lista de productos del carrito y se mueve con ella
        scrollbar = ttk.Scrollbar(self.main_frame, orient=tk.VERTICAL, command=self.carrito_treeview.yview)
        # Configurar la scrollbar para que se mueva junto con la lista de productos del carrito en el eje y (vertical)
        self.carrito_treeview.configure(yscroll=scrollbar.set)
        scrollbar.place(x=1155, y=160, relheight=0.31)

        # Botón para generar el presupuesto
        # El botón llama a la función generar_presupuesto_excel cuando se hace click
        self.generate_button = ttk.Button(self.main_frame, text="Generar Presupuesto", command=self.generar_presupuesto_excel)
        self.generate_button.place(x=560, y=400)

        # Botón para guardar el presupuesto en la base de datos
        self.save_presupuesto_button = ttk.Button(self.main_frame, text="Guardar Presupuesto", command=self.guardar_presupuesto)
        self.save_presupuesto_button.place(x=700, y=400)

        # Botón para generar el remito
        # El botón llama a la función generar_remito_excel cuando se hace click
        self.generate_button = ttk.Button(self.main_frame, text="Generar Remito", command=self.generar_remito_excel)
        self.generate_button.place(x=850, y=400)

        # Botón para guardar el remito en la base de datos
        self.save_remito_button = ttk.Button(self.main_frame, text="Guardar Remito", command=self.guardar_remito)
        self.save_remito_button.place(x=1000, y=400)

    # Funciones para interactuar con la base de datos y la interfaz

    def obtener_nombres_clientes(self):
        clientes = session.query(Clientes.nombre).all()
        return [cliente[0] for cliente in clientes]

    def modificar_precios(self):
        modificar_precios(self.tabla_var, self.precios_anteriores, Categorias, Productos)
        self.update_productos(None)

    def deshacer_ultimo_aumento(self):
        deshacer_ultimo_aumento(self.tabla_var, self.precios_anteriores, Categorias, Productos)
        self.update_productos(None)

    def agregar_fuera_lista(self):
        agregar_fuera_lista(self.carrito, self.producto_var, self.cantidad_fuera_lista_var, self.precio_var)
        self.actualizar_carrito()

    def update_productos(self, event=None):
        update_productos(self.tabla_var, self.productos_tree, Productos, Categorias, event)

    def buscar_producto(self):
        buscar_producto(self.busqueda_var, self.tabla_var, self.productos_tree, Productos, Categorias)

    def agregar_al_carrito(self):
        agregar_al_carrito(self.carrito, self.productos_tree, self.cantidad_var, self.descuento_var)
        self.actualizar_carrito()

    def actualizar_carrito(self):
        actualizar_carrito(self.carrito_treeview, self.carrito)

    def eliminar_del_carrito(self):
        eliminar_del_carrito(self.carrito, self.carrito_treeview)
        self.actualizar_carrito()

    def limpiar_carrito(self):
        # Limpiar la lista de productos del carrito
        self.carrito = []
        # Actualizar la lista de productos del carrito
        self.actualizar_carrito()
        
    def generar_remito_excel(self):
        # Solicitar ubicación para guardar el archivo Excel
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            return
        
        # Cargar la plantilla de Excel para el remito con openpyxl y obtener la hoja activa
        wb = load_workbook('./data/PLANTILLA REMITO.xlsx')
        sheet = wb.active

        # Rellenar los datos del remito

        # Obtener el cliente
        cliente = self.cliente_var.get()

        # Si no se seleccionó un cliente, mostrar un mensaje de error
        if not cliente:
            messagebox.showerror("Error", "Selecciona un cliente.")
            return

        # Si el cliente es "Consumidor Final", no imputamos datos de cliente
        if cliente != "Consumidor Final":
            # Verificar si el cliente está en la base de datos
            cliente_db = session.query(Clientes).filter_by(nombre=cliente).first()
            if cliente_db:
                nombre_cliente = cliente_db.nombre
                cuit_cliente = cliente_db.cuit
                domicilio_cliente = cliente_db.direccion

                # Escribir los datos del cliente en las celdas correspondientes
                sheet.cell(row=6, column=4, value=f"CLIENTE: {nombre_cliente}").font = Font(name='Arial')
                sheet.cell(row=8, column=4, value=f"DOMICILIO: {domicilio_cliente}").font = Font(name='Arial')
                sheet.cell(row=8, column=6, value=f"CUIT: {cuit_cliente}").font = Font(name='Arial')

        # Obtener la fecha actual
        fecha_actual = datetime.date.today().strftime("%d-%m-%Y")

        # Escribir la fecha actual en la celda correspondiente
        sheet.cell(row=4, column=4, value=fecha_actual)

        fila_inicial = 11
        total = 0

        # Iterar sobre los productos en el carrito y hacer los cálculos necesarios
        for item in self.carrito:
            producto, cantidad, _, precio = item
            # Convertir la cantidad y el precio a entero y flotante
            cantidad = int(cantidad)
            precio = float(precio)
            # Calcular el precio total del producto
            precio_total = cantidad * precio

            # Escribir los datos en las celdas correspondientes
            sheet.cell(row=fila_inicial, column=1, value=cantidad)
            sheet.cell(row=fila_inicial, column=2, value=producto)
            sheet.cell(row=fila_inicial, column=7, value=precio)
            sheet.cell(row=fila_inicial, column=8, value=precio_total)

            # Pasar a la siguiente fila
            total += precio_total
            fila_inicial += 1
        
        # Imputar el total
        sheet.cell(row=fila_inicial, column=8, value=total).font = Font(name='Arial', bold=True)

        # Firma del cliente y observaciones
        sheet.cell(row=fila_inicial + 1, column=1, value="FIRMA DEL CLIENTE:").font = Font(name='Arial', bold=True)
        sheet.cell(row=fila_inicial + 1, column=5, value="OBSERVACIONES:").font = Font(name='Arial', bold=True)

        # Copia para la empresa

        if fila_inicial <= 40:
            fila_inicial = 40
            # Fusionar celdas para el título
            sheet.merge_cells(start_row=fila_inicial, start_column=7, end_row=fila_inicial +2, end_column=8)
            # Titulo "REMITO" con fuente 'Arial' de tamaño 20 y negrita
            sheet.cell(row=fila_inicial, column=7, value="REMITO").font = Font(name='Arial', size=20, bold=True, color="4f81bd")
            # Alinear a la derecha
            sheet.cell(row=fila_inicial, column=7).alignment = Alignment(horizontal='right', vertical='top')

            # Etiqueta "Fecha de entrega"
            sheet.cell(row=fila_inicial, column=1, value="Fecha de entrega:")
            # Fecha actual 
            sheet.cell(row=fila_inicial, column=2, value=fecha_actual)

            # Si el cliente no es consumidor final, rellenar
            if cliente != "Consumidor Final":
            # Rellenar los datos del cliente
                sheet.cell(row=fila_inicial + 1, column=1, value=f"NOMBRE: {nombre_cliente}").font = Font(name='Arial')
                sheet.cell(row=fila_inicial + 1, column=3, value=f"CUIT: {cuit_cliente}").font = Font(name='Arial')
                sheet.cell(row=fila_inicial + 3, column=1, value=f"DOMICILIO: {domicilio_cliente}").font = Font(name='Arial')
                sheet.cell(row=fila_inicial + 3, column=3, value="DNI:").font = Font(name='Arial')

            else:
                sheet.cell(row=fila_inicial + 2, column=1, value="CLIENTE: CONSUMIDOR FINAL").font = Font(name='Arial')
            
            # Alinear a la izquierda Fecha de entrega, Cliente y Domicilio
            sheet.cell(row=fila_inicial, column=1).alignment = Alignment(horizontal='left')
            sheet.cell(row=fila_inicial + 1, column=1).alignment = Alignment(horizontal='left')
            sheet.cell(row=fila_inicial + 3, column=1).alignment = Alignment(horizontal='left')

            # Borde para las celdas
            borde_cant = Border(left=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin'))
            
            borde_topbot = Border(top=Side(style='thin'),
                                bottom=Side(style='thin'))
            
            borde_total = Border(top=Side(style='thin'),
                                bottom=Side(style='thin'),
                                right=Side(style='thin'))

            # CANTIDAD, DETALLE, PRECIO UD., TOTAL
            sheet.cell(row=fila_inicial + 5, column=1, value="CANTIDAD").border = borde_cant
            sheet.cell(row=fila_inicial + 5, column=2, value="DETALLE").border = borde_topbot
            for i in range(3, 7):
                sheet.cell(row=fila_inicial + 5, column=i).border = borde_topbot
            sheet.cell(row=fila_inicial + 5, column=7, value="PRECIO UD.").border = borde_topbot
            sheet.cell(row=fila_inicial + 5, column=8, value="TOTAL").border = borde_total

            # Aplicar negrita a la fila de los encabezados
            sheet.cell(row=fila_inicial + 5, column=1).font = Font(name='Arial', bold=True)
            sheet.cell(row=fila_inicial + 5, column=2).font = Font(name='Arial', bold=True)
            sheet.cell(row=fila_inicial + 5, column=7).font = Font(name='Arial', bold=True)
            sheet.cell(row=fila_inicial + 5, column=8).font = Font(name='Arial', bold=True)

            fila_inicial += 6

            # Imputar los datos del carrito
            for item in self.carrito:
                producto, cantidad, _, precio = item
                # Convertir la cantidad y el precio a entero y flotante
                cantidad = int(cantidad)
                precio = float(precio)
                # Calcular el precio total del producto
                precio_total = cantidad * precio

                # Escribir los datos en las celdas correspondientes

                sheet.cell(row=fila_inicial, column=1, value=cantidad)
                sheet.cell(row=fila_inicial, column=2, value=producto)
                sheet.cell(row=fila_inicial, column=7, value=precio)
                sheet.cell(row=fila_inicial, column=8, value=precio_total)

                # Pasar a la siguiente fila
                fila_inicial += 1
            
            # Imputar el total
            sheet.cell(row=fila_inicial, column=8, value=total).font = Font(name='Arial', bold=True)    

            # Firma del cliente y observaciones
            sheet.cell(row=fila_inicial + 1, column=1, value="FIRMA DEL CLIENTE:").font = Font(name='Arial', bold=True)     
            sheet.cell(row=fila_inicial + 1, column=5, value="OBSERVACIONES:").font = Font(name='Arial', bold=True)

        # Guardar el archivo Excel
        wb.save(file_path)
        messagebox.showinfo("Éxito", f"Remito generado en {file_path}")

        # Abrir el archivo automáticamente
        try:
            os.startfile(file_path)

        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir el archivo: {str(e)}")

    def guardar_remito(self):
        guardar_remito(self.cliente_var, self.carrito, session, Clientes, Remitos, DetallesRemitos)
        
      

    def generar_presupuesto_excel(self):
        # Solicitar ubicación para guardar el archivo Excel
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            return

        # Cargar la plantilla de Excel para el presupuesto con openpyxl y obtener la hoja activa
        wb = load_workbook('./data/PLANTILLA PRESUPUESTO.xlsx')
        sheet = wb.active

        # Obtener el cliente
        cliente = self.cliente_var.get()

        # Si no se seleccionó un cliente, mostrar un mensaje de error
        if not cliente:
            messagebox.showerror("Error", "Selecciona un cliente.")
            return

        # Obtener la fecha actual
        fecha_actual = datetime.date.today().strftime("%d-%m-%Y")

        # Rellenar los datos generales del presupuesto
        # Fecha
        sheet.cell(row=3, column=6, value=fecha_actual)

        # Si el cliente es "Consumidor Final", no imputamos datos de cliente
        if cliente != "Consumidor Final":
            # Verificar si el cliente está en la base de datos
            cliente_db = session.query(Clientes).filter_by(nombre=cliente).first()
            if cliente_db:
                nombre_cliente = cliente_db.nombre
                cuit_cliente = cliente_db.cuit
                domicilio_cliente = cliente_db.direccion

                # Escribir los datos del cliente en las celdas correspondientes
                sheet.cell(row=5, column=6, value=nombre_cliente)
                sheet.cell(row=6, column=6, value=cuit_cliente)
                sheet.cell(row=7, column=6, value=domicilio_cliente)

        else:
            # Si el cliente es "Consumidor Final", escribir "Consumidor Final" en la celda correspondiente
            sheet.cell(row=5, column=6, value="Consumidor Final")

        # Estilo de borde y relleno para las celdas aplicado con openpyxl
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
        blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

        # Rellenar la plantilla con los datos del presupuesto
        fila_inicial = 11  # Fila donde comienzan los productos en la plantilla
        subtotal = 0

        # Iterar sobre los productos en el carrito y hacer los cálculos necesarios
        for item in self.carrito:
            # Desempaquetar los datos de la lista del carrito en variables separadas 
            producto, cantidad, descuento, precio = item
            # Convertir la cantidad y el precio a entero y flotante
            cantidad = int(cantidad)
            precio = float(precio)
            # Calcular el precio total con descuento y sin IVA del producto
            precio_total = cantidad * precio * (1 - descuento / 100)
            precio_sin_iva = precio_total / 1.21
            # Convertir el descuento a porcentaje
            descuento_porcentaje = float(descuento) / 100

            # Escribir los datos en las celdas correspondientes
            sheet.cell(row=fila_inicial, column=2, value=cantidad)
            sheet.cell(row=fila_inicial, column=3, value=producto)
            sheet.cell(row=fila_inicial, column=4, value=precio)
            sheet.cell(row=fila_inicial, column=5, value=descuento_porcentaje).number_format = '0.00%'
            sheet.cell(row=fila_inicial, column=6, value=precio_total)

            # Aplicar borde a las celdas de la fila actual
            for col in range(2, 7):
                sheet.cell(row=fila_inicial, column=col).border = thin_border

            # Calcular el subtotal del presupuesto con los precios sin IVA
            subtotal += precio_sin_iva
            # Pasar a la siguiente fila
            fila_inicial += 1

        # Calcular los totales
        iva = subtotal * 0.21
        total = subtotal + iva

        # Escribir los totales en las celdas correspondientes
        sheet.cell(row=fila_inicial, column=5, value="Subtotal").font = Font(bold=True)
        sheet.cell(row=fila_inicial, column=6, value=subtotal)

        sheet.cell(row=fila_inicial + 1, column=5, value="IVA").font = Font(bold=True)
        sheet.cell(row=fila_inicial + 1, column=6, value=iva)

        sheet.cell(row=fila_inicial + 2, column=5, value="Total").font = Font(bold=True)
        sheet.cell(row=fila_inicial + 2, column=6, value=total)

        # Gracias por su confianza en la tercer columna
        sheet.cell(row=fila_inicial + 2, column=3, value="GRACIAS POR SU CONFIANZA").font = Font(bold=True)

        # Aplicar borde y relleno a las celdas de subtotal, iva y total
        for row in range(fila_inicial, fila_inicial + 3):
            for col in range(5, 7):
                cell = sheet.cell(row=row, column=col)
                cell.border = thin_border
                # Aplicar relleno azul a las celdas de subtotal, iva y total
                # cell.fill = blue_fill

        # Copia para la empresa

        if fila_inicial <= 40:
            fila_inicial = 40
            # Fusionar celdas para el título
            sheet.merge_cells(start_row=fila_inicial, start_column=5, end_row=fila_inicial +2, end_column=6)
            # Titulo "PRESUPUESTO" con fuente 'Arial' de tamaño 20 y negrita
            sheet.cell(row=fila_inicial, column=5, value="PRESUPUESTO").font = Font(name='Arial', size=27, bold=True, color="8db3e2")
            # Alinear a la derecha y arriba
            sheet.cell(row=fila_inicial, column=5).alignment = Alignment(horizontal='right', vertical='top')

            # Etiqueta "Fecha"
            sheet.cell(row=fila_inicial, column=4, value=f"Fecha: {fecha_actual}")

            # Guardar el nombre del cliente en una variable
            nombre_cliente = self.cliente_var.get()

            # Si el cliente no es "Consumidor Final", rellenar los datos del cliente
            if nombre_cliente != "Consumidor Final":
                # Etiqueta "CLIENTE" y nombre del cliente alineado a la izquierda
                sheet.cell(row=fila_inicial, column=2, value=f"CLIENTE: {nombre_cliente}").alignment = Alignment(horizontal='left')

                # Etiqueta "Domicilio" y domicilio del cliente
                sheet.cell(row=fila_inicial + 1, column=2, value=f"Domicilio: {domicilio_cliente}").alignment = Alignment(horizontal='left')

                # Etiqueta "CUIT" y CUIT del cliente
                sheet.cell(row=fila_inicial + 2, column=2, value=f"CUIT: {cuit_cliente}").alignment = Alignment(horizontal='left')

                # Etiqueta "DNI" y espacio para completar
                sheet.cell(row=fila_inicial + 3, column=2, value="DNI:").alignment = Alignment(horizontal='left')

            else:
                # Si el cliente es "Consumidor Final", escribir "Consumidor Final" en la celda correspondiente
                sheet.cell(row=fila_inicial + 1, column=2, value="CLIENTE: Consumidor Final").alignment = Alignment(horizontal='left')

            # Borde para las celdas
            borde = Border(left=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin'),
                                right=Side(style='thin'))
            
            # CANTIDAD, DETALLE, PRECIO UD., DESCUENTO, IMPORTE C/IVA
            sheet.cell(row=fila_inicial + 4, column=2, value="CANTIDAD").border = borde
            sheet.cell(row=fila_inicial + 4, column=3, value="DETALLE").border = borde
            sheet.cell(row=fila_inicial + 4, column=4, value="PRECIO UD.").border = borde
            sheet.cell(row=fila_inicial + 4, column=5, value="DESCUENTO").border = borde
            sheet.cell(row=fila_inicial + 4, column=6, value="IMPORTE C/IVA").border = borde

            # Color
            color_fondo = PatternFill(start_color="dbe5f1", end_color="dbe5f1", fill_type="solid")

            # Aplicar color de fondo dbe5f1 a la fila de los encabezados
            for col in range(2, 7):
                sheet.cell(row=fila_inicial + 4, column=col).fill = color_fondo

                  # Rellenar la plantilla con los datos del presupuesto
        fila_inicial += 5  # Fila donde comienzan los productos en la plantilla
        subtotal = 0

        # Iterar sobre los productos en el carrito y hacer los cálculos necesarios
        for item in self.carrito:
            # Desempaquetar los datos de la lista del carrito en variables separadas 
            producto, cantidad, descuento, precio = item
            # Convertir la cantidad y el precio a entero y flotante
            cantidad = int(cantidad)
            precio = float(precio)
            # Calcular el precio total con descuento y sin IVA del producto
            precio_total = cantidad * precio * (1 - descuento / 100)
            precio_sin_iva = precio_total / 1.21
            # Convertir el descuento a porcentaje
            descuento_porcentaje = float(descuento) / 100

            # Escribir los datos en las celdas correspondientes
            sheet.cell(row=fila_inicial, column=2, value=cantidad)
            sheet.cell(row=fila_inicial, column=3, value=producto)
            sheet.cell(row=fila_inicial, column=4, value=precio)
            sheet.cell(row=fila_inicial, column=5, value=descuento_porcentaje).number_format = '0.00%'
            sheet.cell(row=fila_inicial, column=6, value=precio_total)

            # Aplicar borde a las celdas de la fila actual
            for col in range(2, 7):
                sheet.cell(row=fila_inicial, column=col).border = borde

            # Calcular el subtotal del presupuesto con los precios sin IVA
            subtotal += precio_sin_iva
            # Pasar a la siguiente fila
            fila_inicial += 1

        # Calcular los totales
        iva = subtotal * 0.21
        total = subtotal + iva

        # Escribir los totales en las celdas correspondientes
        sheet.cell(row=fila_inicial, column=5, value="Subtotal").font = Font(bold=True)
        sheet.cell(row=fila_inicial, column=6, value=subtotal)

        sheet.cell(row=fila_inicial + 1, column=5, value="IVA").font = Font(bold=True)
        sheet.cell(row=fila_inicial + 1, column=6, value=iva)

        sheet.cell(row=fila_inicial + 2, column=5, value="Total").font = Font(bold=True)
        sheet.cell(row=fila_inicial + 2, column=6, value=total)

        # Gracias por su confianza en la tercer columna
        sheet.cell(row=fila_inicial + 2, column=3, value="GRACIAS POR SU CONFIANZA").font = Font(bold=True)

        # Aplicar borde y relleno a las celdas de subtotal, iva y total
        for row in range(fila_inicial, fila_inicial + 3):
            for col in range(5, 7):
                cell = sheet.cell(row=row, column=col)
                cell.border = borde

        # Guardar el archivo Excel
        wb.save(file_path)
        messagebox.showinfo("Éxito", f"Presupuesto generado en {file_path}")

        # Abrir el archivo automáticamente
        try:
            os.startfile(file_path)

        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir el archivo: {str(e)}")

    def guardar_presupuesto(self):
        try:
            # Obtener el nombre del cliente seleccionado
            nombre_cliente = self.cliente_var.get()

            # Buscar el cliente en la base de datos
            cliente = session.query(Clientes).filter_by(nombre=nombre_cliente).first()

            if not cliente:
                messagebox.showerror("Error", "Cliente no encontrado en la base de datos.")
                return

            # Obtener la fecha actual
            fecha_actual = datetime.date.today().strftime("%d-%m-%Y")

            # Crear un nuevo presupuesto
            nuevo_presupuesto = Presupuestos(id_cliente=cliente.id, fecha=fecha_actual, total=0)

            # Agregar el presupuesto a la base de datos
            session.add(nuevo_presupuesto)
            session.commit()

            total_presupuesto = 0

            # Iterar sobre los elementos del carrito para agregarlos a los detalles del presupuesto
            for producto, cantidad, descuento, precio in self.carrito:
                cantidad = int(cantidad)
                precio = float(precio)
                descuento = float(descuento)
                total = cantidad * precio * (1 - descuento / 100)
                detalle = DetallesPresupuestos(
                    id_presupuesto=nuevo_presupuesto.id,
                    producto=producto,
                    cantidad=cantidad,
                    precio_unitario=precio,
                    descuento=descuento,
                    total=total
                )
                total_presupuesto += total
                session.add(detalle)

            # Actualizar el total del presupuesto
            nuevo_presupuesto.total = total_presupuesto
            session.commit()

            messagebox.showinfo("Éxito", "El presupuesto se ha guardado correctamente en la base de datos.")
        
        except Exception as e:
            session.rollback()
            messagebox.showerror("Error", f"Ocurrió un error al guardar el presupuesto: {str(e)}")

    def mostrar_clientes(self):
        # Limpiar el main_frame antes de agregar nuevos widgets
        for widget in self.main_frame.winfo_children():
            widget.destroy()

        # Crear una instancia de ClientesApp y agregar sus widgets al main_frame
        self.clientes_app = ClientesApp(self.main_frame)         

# Función principal para ejecutar la aplicación
# Si el script se ejecuta directamente, se crea una instancia de la clase PresupuestoApp y se llama al método mainloop
if __name__ == "__main__":
    app = RemitosApp()
    app.mainloop()
