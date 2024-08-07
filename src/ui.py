import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
from reportlab.lib.pagesizes import letter
from sqlalchemy.orm import sessionmaker
from database import engine, Vasser, Capea, Polietileno, Peirano, Latyn, Fusiogas, Chicote, H3, CañosPVC, PiezasPVC, Sigas, PPRosca, Awaduck, Amancofusion, Rotoplas
from openpyxl import load_workbook
import datetime
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
import os

# Sesión de SQLAlchemy para interactuar con la base de datos
Session = sessionmaker(bind=engine)
session = Session()

# Diccionario para mapear nombres de tablas a clases
tabla_clase_mapping = {
    'VASSER': Vasser,
    'CAPEA': Capea,
    'POLIETILENO': Polietileno,
    'PEIRANO': Peirano,
    'LATYN': Latyn,
    'FUSIOGAS': Fusiogas,
    'CHICOTE': Chicote,
    'H3': H3,
    'CAÑOS PVC': CañosPVC,
    'PIEZAS PVC Y LOSUNG': PiezasPVC,
    'SIGAS': Sigas,
    'PP ROSCA': PPRosca,
    'AWADUCK': Awaduck,
    'AMANCO FUSION': Amancofusion,
    'ROTOPLAS': Rotoplas
}

# Clase para la interfaz
class PresupuestoApp(tk.Tk):
    def __init__(self):
        # Inicializar la ventana principal
        super().__init__()
        # Título de la ventana
        self.title("Sistema de Presupuestos")
        # Geometría de la ventana
        self.geometry("1280x720")
        # Crear los widgets con la función create_widgets
        self.create_widgets()
        # Lista para almacenar los productos del carrito
        self.carrito = []

    def create_widgets(self):
        # Menú lateral
        # Frame para el menú lateral con un ancho de 200 y color de fondo gris
        self.menu_lateral = tk.Frame(self, width=200, bg='gray')
        # Empaquetar el frame en la ventana principal a la izquierda y que se expanda en el eje y
        self.menu_lateral.pack(side='left', fill='y')

        # Botón para mostrar presupuestos
        # ttk.Button es un botón con un estilo mejorado
        # El botón llama a la función mostrar_presupuestos cuando se hace click 
        self.presupuestos_button = ttk.Button(self.menu_lateral, text="Presupuestos", command=self.mostrar_presupuestos)
        # Empaquetar el botón en el menú lateral
        self.presupuestos_button.pack(padx=10, pady=10)

        # Botón para mostrar clientes
        self.clientes_button = ttk.Button(self.menu_lateral, text="Clientes", command=self.mostrar_clientes)
        self.clientes_button.pack(padx=10, pady=10)

        # Frame principal
        self.main_frame = tk.Frame(self)
        # Empaquetar el frame principal en la ventana principal
        self.main_frame.pack(side='right', fill='both', expand=True)

    def mostrar_presupuestos(self):
        # Limpiar el main_frame antes de agregar nuevos widgets para evitar superposiciones
        for widget in self.main_frame.winfo_children():
            # Destruir el widget
            widget.destroy()

        # Menú de selección de tablas (listas de precios)
        self.tabla_label = ttk.Label(self.main_frame, text="Seleccioná una lista:")
        # Empaquetar la etiqueta en el main_frame con place
        self.tabla_label.place(x=10, y=10)

        # Combobox para seleccionar la tabla

        # StringVar para almacenar el valor seleccionado en el Combobox
        self.tabla_var = tk.StringVar()
        # Crear el Combobox con las opciones de las tablas, textvariable es la variable que almacena el valor seleccionado 
        self.tabla_combobox = ttk.Combobox(self.main_frame, textvariable=self.tabla_var)
        # Asignar las opciones del Combobox a las claves del diccionario tabla_clase_mapping
        self.tabla_combobox['values'] = list(tabla_clase_mapping.keys())
        # Seleccionar la primera opción por defecto
        self.tabla_combobox.bind("<<ComboboxSelected>>", self.update_productos)

        self.tabla_combobox.place(x=135, y=10)

        # Botón para aumentar precios
        # El botón llama a la función aumentar_precios cuando se hace click
        self.aumentar_precios_button = ttk.Button(self.main_frame, text="Aumentar Precios", command=self.aumentar_precios)
        self.aumentar_precios_button.place(x=300, y=8)

        # Etiqueta de búsqueda
        self.busqueda_label = ttk.Label(self.main_frame, text="Buscar Producto:")
        self.busqueda_label.place(x=10, y=40)

        # Campo de búsqueda
        # StringVar para almacenar el valor ingresado en el campo de búsqueda 
        self.busqueda_var = tk.StringVar()
        # Entry para ingresar el término de búsqueda, textvariable es la variable que almacena el valor ingresado
        self.busqueda_entry = ttk.Entry(self.main_frame, textvariable=self.busqueda_var)
        self.busqueda_entry.place(x=135, y=40)

        # Botón para buscar productos
        # El botón llama a la función buscar_producto cuando se hace click
        self.buscar_button = ttk.Button(self.main_frame, text="Buscar", command=self.buscar_producto)
        self.buscar_button.place(x=300, y=38)

        # Etiqueta de cantidad
        self.cantidad_label = ttk.Label(self.main_frame, text="Cantidad:")
        self.cantidad_label.place(x=10, y=70)

        # Campo de cantidad
        # IntVar para almacenar el valor ingresado en el campo de cantidad 
        self.cantidad_var = tk.IntVar()
        # Entry para ingresar la cantidad, textvariable es la variable que almacena el valor ingresado
        self.cantidad_entry = ttk.Entry(self.main_frame, textvariable=self.cantidad_var)
        self.cantidad_entry.place(x=135, y=70)

        # Etiqueta de descuento
        self.descuento_label = ttk.Label(self.main_frame, text="Descuento (%):")
        self.descuento_label.place(x=10, y=100)

        # Campo de descuento
        # DoubleVar para almacenar el valor ingresado en el campo de descuento
        self.descuento_var = tk.DoubleVar()
        # Entry para ingresar el descuento, textvariable es la variable que almacena el valor ingresado
        self.descuento_entry = ttk.Entry(self.main_frame, textvariable=self.descuento_var)
        self.descuento_entry.place(x=135, y=100)

        # Botón para agregar al carrito
        # El botón llama a la función agregar_al_carrito cuando se hace click
        self.add_button = ttk.Button(self.main_frame, text="Agregar al Carrito", command=self.agregar_al_carrito)
        self.add_button.place(x=135, y=130)

        # Tabla de productos
        # Treeview con las columnas Producto y Precio
        self.productos_tree = ttk.Treeview(self.main_frame, columns=('Codigo', 'Linea', 'Producto', 'Precio'), show='headings')
        # Encabezados de las columnas
        self.productos_tree.heading('Codigo', text='Codigo')
        self.productos_tree.heading('Linea', text='Linea')
        self.productos_tree.heading('Producto', text='Producto')
        self.productos_tree.heading('Precio', text='Precio')
        # Ancho de las columnas
        self.productos_tree.column('Codigo', width=50)
        self.productos_tree.column('Linea', width=70)
        self.productos_tree.column('Producto', width=300)
        self.productos_tree.column('Precio', width=100)
        self.productos_tree.place(x=10, y=160)

        # Scrollbar para la tabla de productos
        # Scrollbar en el eje vertical que se conecta con la tabla de productos y se mueve con ella
        scrollbar = ttk.Scrollbar(self.main_frame, orient=tk.VERTICAL, command=self.productos_tree.yview)
        # Configurar la scrollbar para que se mueva junto con la tabla de productos en el eje y (vertical)
        self.productos_tree.configure(yscroll=scrollbar.set)
        scrollbar.place(x=532, y=160, relheight=0.31)

        # Carrito

        # Agregar productos fuera de lista
        # Etiqueta de productos fuera de lista
        self.productos_fuera_lista_label = ttk.Label(self.main_frame, text="Productos Fuera de Lista")
        self.productos_fuera_lista_label.place(x=750, y=10)

        # Etiqueta para nombre del producto
        self.producto_label = ttk.Label(self.main_frame, text="Producto:")
        self.producto_label.place(x=750, y=40)

        # Campo para ingresar el nombre del producto
        self.producto_var = tk.StringVar()
        self.producto_entry = ttk.Entry(self.main_frame, textvariable=self.producto_var)
        self.producto_entry.place(x=815, y=40)

        # Etiqueta para cantidad del producto
        self.cantidad_fuera_lista_label = ttk.Label(self.main_frame, text="Cantidad:")
        self.cantidad_fuera_lista_label.place(x=750, y=70)

        # Campo para ingresar la cantidad del producto
        self.cantidad_fuera_lista_var = tk.IntVar()
        self.cantidad_fuera_lista_entry = ttk.Entry(self.main_frame, textvariable=self.cantidad_fuera_lista_var)
        self.cantidad_fuera_lista_entry.place(x=815, y=70)

        # Etiqueta para precio del producto
        self.precio_label = ttk.Label(self.main_frame, text="Precio:")
        self.precio_label.place(x=750, y=100)

        # Campo para ingresar el precio del producto
        self.precio_var = tk.DoubleVar()
        self.precio_entry = ttk.Entry(self.main_frame, textvariable=self.precio_var)
        self.precio_entry.place(x=815, y=100)

        # Botón para agregar productos fuera de lista
        self.add_fuera_lista_button = ttk.Button(self.main_frame, text="Agregar al Carrito", command=self.agregar_fuera_lista)
        self.add_fuera_lista_button.place(x=750, y=130)

        # Treeview para mostrar los productos del carrito
        self.carrito_treeview = ttk.Treeview(self.main_frame, columns=('Producto', 'Cantidad', 'Descuento', 'Precio'), show='headings')
        self.carrito_treeview.heading('Producto', text='Producto')
        self.carrito_treeview.heading('Cantidad', text='Cantidad')
        self.carrito_treeview.heading('Descuento', text='Descuento')
        self.carrito_treeview.heading('Precio', text='Precio')
        self.carrito_treeview.column('Producto', width=350)
        self.carrito_treeview.column('Cantidad', width=75)
        self.carrito_treeview.column('Descuento', width=75)
        self.carrito_treeview.column('Precio', width=100)
        self.carrito_treeview.place(x=552, y=160)

        self.actualizar_carrito()

        # Scrollbar para la lista de productos del carrito
        # Scrollbar en el eje vertical que se conecta con la lista de productos del carrito y se mueve con ella
        scrollbar = ttk.Scrollbar(self.main_frame, orient=tk.VERTICAL, command=self.carrito_treeview.yview)
        # Configurar la scrollbar para que se mueva junto con la lista de productos del carrito en el eje y (vertical)
        self.carrito_treeview.configure(yscroll=scrollbar.set)
        scrollbar.place(x=1155, y=160, relheight=0.31)

        # Botón para borrar del carrito
        # El botón llama a la función eliminar_del_carrito cuando se hace click
        self.delete_button = ttk.Button(self.main_frame, text="Eliminar del Carrito", command=self.eliminar_del_carrito)
        self.delete_button.place(x=585, y=400)

        # Botón para generar el presupuesto
        # El botón llama a la función generar_presupuesto_excel cuando se hace click
        self.generate_button = ttk.Button(self.main_frame, text="Generar Presupuesto", command=self.generar_presupuesto_excel)
        self.generate_button.place(x=780, y=400)

        # Botón para generar el remito
        # El botón llama a la función generar_remito_excel cuando se hace click
        self.generate_button = ttk.Button(self.main_frame, text="Generar Remito", command=self.generar_remito_excel)
        self.generate_button.place(x=1000, y=400)


    def mostrar_clientes(self):
        # Limpiar el main_frame antes de agregar nuevos widgets
        for widget in self.main_frame.winfo_children():
            widget.destroy()

        # Label de seccion en desarrollo
        self.seccion_label = ttk.Label(self.main_frame, text="Sección en desarrollo")
        self.seccion_label.place(x=10, y=10)
        

    def aumentar_precios(self):
        # Obtener la tabla seleccionada
        tabla = self.tabla_var.get()
        if not tabla:
            messagebox.showerror("Error", "Selecciona una lista.")
            return

        # Solicitar el porcentaje de aumento al usuario
        aumento = simpledialog.askfloat("Aumentar Precios", "Ingrese el porcentaje de aumento (por ejemplo, 10 para un 10%):")
        # Si el usuario cancela la operación o ingresa un valor inválido, mostrar un mensaje de error
        if aumento is None or aumento <= 0:
            messagebox.showerror("Error", "Ingrese un porcentaje de aumento válido.")
            return

        # Obtener la clase correspondiente a la tabla seleccionada por el usuario del diccionario de mapeo
        clase = tabla_clase_mapping[tabla]
        # Obtener todos los productos de la tabla seleccionada por el usuario
        productos = session.query(clase).all()
        # Aumentar los precios de los productos en la tabla seleccionada
        for producto in productos:
            producto.precio *= (1 + aumento / 100)
        session.commit()

        # Actualizar la tabla de productos, no es necesario pasar el evento como argumento
        self.update_productos(None)
        messagebox.showinfo("Éxito", f"Los precios de {tabla} han sido aumentados en un {aumento}%.")

    def agregar_fuera_lista(self):
        # Obtener el nombre y precio del producto ingresados por el usuario
        producto = self.producto_var.get()
        cantidad = self.cantidad_fuera_lista_var.get()
        precio = self.precio_var.get()

        # Validar que el producto y el precio no estén vacíos
        if not producto or not precio:
            messagebox.showerror("Error", "Ingresa un producto y un precio.")
            return

        # Agregar el producto a la lista de productos del carrito
        self.carrito.append((producto, cantidad, 0, precio))
        # Actualizar la lista de productos del carrito
        self.actualizar_carrito()

    def update_productos(self, event):
        # Obtener la tabla seleccionada por el usuario
        tabla = self.tabla_var.get()
        # Obtener la clase correspondiente a la tabla seleccionada por el usuario del diccionario de mapeo
        clase = tabla_clase_mapping[tabla]

        # Obtener el término de búsqueda y convertirlo a minúsculas 
        search_term = self.busqueda_var.get().lower()

        # Obtener todos los productos de la tabla seleccionada 
        productos = session.query(clase).all()
        # Limpiar la tabla de productos
        self.productos_tree.delete(*self.productos_tree.get_children())

        # Verificar si la clase tiene un atributo 'codigo' y 'linea' para mostrar en la tabla
        if hasattr(clase, 'codigo') and hasattr(clase, 'linea'):
            for producto in productos:
                if search_term in producto.producto.lower():
                    self.productos_tree.insert('', 'end', values=(producto.codigo, producto.linea, producto.producto, producto.precio))
        else:
            for producto in productos:
                if search_term in producto.producto.lower():
                    self.productos_tree.insert('', 'end', values=(None,None,producto.producto, producto.precio))

    def buscar_producto(self):
        # Obtener el término de búsqueda y convertirlo a minúsculas
        search_term = self.busqueda_var.get().lower()
        # Obtener la tabla seleccionada por el usuario
        tabla = self.tabla_var.get()
        # Obtener la clase correspondiente a la tabla seleccionada por el usuario del diccionario de mapeo
        clase = tabla_clase_mapping[tabla]

        # Obtener todos los productos de la tabla seleccionada por el usuario con el método query de la sesión, que recibe la clase (tabla) como argumento
        productos = session.query(clase).all()
        # Limpiar la tabla de productos antes de insertar aquellos que coincidan con el término de búsqueda
        self.productos_tree.delete(*self.productos_tree.get_children())

        # Verificar si la clase tiene un atributo 'codigo' y 'linea' para mostrar en la tabla
        if hasattr(clase, 'codigo') and hasattr(clase, 'linea'):
            for producto in productos:
                # Si la búsqueda está contenida en el nombre del producto, insertar el producto en la tabla
                if search_term in producto.producto.lower():
                    self.productos_tree.insert('', 'end', values=(producto.codigo, producto.linea, producto.producto, producto.precio))
                # Si la búsqueda está contenida en el código del producto, insertar el producto en la tabla
                elif search_term in producto.codigo.lower():
                    self.productos_tree.insert('', 'end', values=(producto.codigo, producto.linea, producto.producto, producto.precio))
                # Si la búsqueda está contenida en la línea del producto, insertar el producto en la tabla
                elif search_term in producto.linea.lower():
                    self.productos_tree.insert('', 'end', values=(producto.codigo, producto.linea, producto.producto, producto.precio))
        else:
            for producto in productos:
                if search_term in producto.producto.lower():
                    self.productos_tree.insert('', 'end', values=(None,None,producto.producto, producto.precio))

    def agregar_al_carrito(self):
        # Obtener el producto seleccionado en la tabla de productos
        selected_item = self.productos_tree.selection()

        if not selected_item:
            messagebox.showerror("Error", "Selecciona un producto.")
            return

        # Obtener el nombre y precio del producto seleccionado a través del índice
        item = self.productos_tree.item(selected_item)
        producto_nombre = item['values'][2]
        precio = item['values'][3]
        # Obtener la cantidad y descuento ingresados por el usuario con el método get
        cantidad = self.cantidad_var.get()
        descuento = self.descuento_var.get()

        if cantidad <= 0:
            messagebox.showerror("Error", "Ingresa una cantidad válida.")
            return

        # Agregar el producto al carrito y actualizar la lista de productos
        self.carrito.append((producto_nombre, cantidad, descuento, precio))
        self.actualizar_carrito()

    def actualizar_carrito(self):
        # Limpiar la lista de productos del carrito antes de insertar los productos actuales
        self.carrito_treeview.delete(*self.carrito_treeview.get_children())
        # Iterar sobre la lista de productos del carrito y agregarlos al treeview
        for item in self.carrito:
            producto, cantidad, descuento, precio = item
            self.carrito_treeview.insert('', 'end', values=(producto, cantidad, descuento, precio))

    def eliminar_del_carrito(self):
        # Obtener el índice del producto seleccionado en el treeview del carrito
        seleccion = self.carrito_treeview.selection()
        if not seleccion:
            messagebox.showerror("Error", "Selecciona un producto del carrito.")
            return
        
        index = self.carrito_treeview.index(seleccion)

        del self.carrito[index]

        self.actualizar_carrito()
        
    def generar_remito_excel(self):
        # Solicitar ubicación para guardar el archivo Excel
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            return
        
        # Cargar la plantilla de Excel para el remito con openpyxl y obtener la hoja activa
        wb = load_workbook('./data/PLANTILLA REMITO.xlsx')
        sheet = wb.active

        # Obtener la fecha actual
        fecha_actual = datetime.date.today().strftime("%d-%m-%Y")

        # Rellenar los datos del remito
        # Fecha
        sheet.cell(row=4, column=4, value=fecha_actual)

        fila_inicial = 11
        total = 0

        # Iterar sobre los productos en el carrito y hacer los cálculos necesarios
        for item in self.carrito:
            producto, cantidad, _, precio = item
            # Convertir la cantidad y el precio a entero y flotante
            cantidad = int(cantidad)
            precio = float(precio)
            # Calcular el precio total del producto
            precio_total = cantidad * precio

            # Escribir los datos en las celdas correspondientes
            sheet.cell(row=fila_inicial, column=1, value=cantidad)
            sheet.cell(row=fila_inicial, column=2, value=producto)
            sheet.cell(row=fila_inicial, column=7, value=precio)
            sheet.cell(row=fila_inicial, column=8, value=precio_total)

            # Pasar a la siguiente fila
            total += precio_total
            fila_inicial += 1
        
        # Imputar el total
        sheet.cell(row=fila_inicial, column=8, value=total)

        # Firma del cliente y observaciones
        sheet.cell(row=fila_inicial + 1, column=1, value="FIRMA DEL CLIENTE:").font = Font(name='Arial', bold=True)
        sheet.cell(row=fila_inicial + 1, column=5, value="OBSERVACIONES:").font = Font(name='Arial', bold=True)

        # Copia para la empresa

        if fila_inicial <= 40:
            fila_inicial = 40
            # Fusionar celdas para el título
            sheet.merge_cells(start_row=fila_inicial, start_column=7, end_row=fila_inicial +2, end_column=8)
            # Titulo "REMITO" con fuente 'Arial' de tamaño 20 y negrita
            sheet.cell(row=fila_inicial, column=7, value="REMITO").font = Font(name='Arial', size=20, bold=True, color="4f81bd")
            # Alinear a la derecha
            sheet.cell(row=fila_inicial, column=7).alignment = Alignment(horizontal='right', vertical='top')

            # Etiqueta "Fecha de entrega"
            sheet.cell(row=fila_inicial, column=1, value="Fecha de entrega:")
            # Fecha actual 
            sheet.cell(row=fila_inicial, column=2, value=fecha_actual)

            # Etiqueta "CLIENTE" 
            sheet.cell(row=fila_inicial + 1, column=1, value="CLIENTE:")

            # Etiqueta "DNI"
            sheet.cell(row=fila_inicial + 1, column=3, value="DNI:")

            # Etiqueta "Domicilio"
            sheet.cell(row=fila_inicial + 3, column=1, value="DOMICILIO:")

            # Alinear a la izquierda Fecha de entrega, Cliente y Domicilio
            sheet.cell(row=fila_inicial, column=1).alignment = Alignment(horizontal='left')
            sheet.cell(row=fila_inicial + 1, column=1).alignment = Alignment(horizontal='left')
            sheet.cell(row=fila_inicial + 3, column=1).alignment = Alignment(horizontal='left')

            # Etiqueta "CUIT"
            sheet.cell(row=fila_inicial + 3, column=3, value="CUIT:")

            # Borde para las celdas
            borde_cant = Border(left=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin'))
            
            borde_topbot = Border(top=Side(style='thin'),
                                bottom=Side(style='thin'))
            
            borde_total = Border(top=Side(style='thin'),
                                bottom=Side(style='thin'),
                                right=Side(style='thin'))

            # CANTIDAD, DETALLE, PRECIO UD., TOTAL
            sheet.cell(row=fila_inicial + 5, column=1, value="CANTIDAD").border = borde_cant
            sheet.cell(row=fila_inicial + 5, column=2, value="DETALLE").border = borde_topbot
            for i in range(3, 7):
                sheet.cell(row=fila_inicial + 5, column=i).border = borde_topbot
            sheet.cell(row=fila_inicial + 5, column=7, value="PRECIO UD.").border = borde_topbot
            sheet.cell(row=fila_inicial + 5, column=8, value="TOTAL").border = borde_total

            # Aplicar negrita a la fila de los encabezados
            sheet.cell(row=fila_inicial + 5, column=1).font = Font(name='Arial', bold=True)
            sheet.cell(row=fila_inicial + 5, column=2).font = Font(name='Arial', bold=True)
            sheet.cell(row=fila_inicial + 5, column=7).font = Font(name='Arial', bold=True)
            sheet.cell(row=fila_inicial + 5, column=8).font = Font(name='Arial', bold=True)

            fila_inicial += 6

            # Imputar los datos del carrito
            for item in self.carrito:
                producto, cantidad, _, precio = item
                # Convertir la cantidad y el precio a entero y flotante
                cantidad = int(cantidad)
                precio = float(precio)
                # Calcular el precio total del producto
                precio_total = cantidad * precio

                # Escribir los datos en las celdas correspondientes

                sheet.cell(row=fila_inicial, column=1, value=cantidad)
                sheet.cell(row=fila_inicial, column=2, value=producto)
                sheet.cell(row=fila_inicial, column=7, value=precio)
                sheet.cell(row=fila_inicial, column=8, value=precio_total)

                # Pasar a la siguiente fila
                fila_inicial += 1
            
            # Imputar el total
            sheet.cell(row=fila_inicial, column=8, value=total)       

            # Firma del cliente y observaciones
            sheet.cell(row=fila_inicial + 1, column=1, value="FIRMA DEL CLIENTE:").font = Font(name='Arial', bold=True)     
            sheet.cell(row=fila_inicial + 1, column=5, value="OBSERVACIONES:").font = Font(name='Arial', bold=True)

        # Guardar el archivo Excel
        wb.save(file_path)
        messagebox.showinfo("Éxito", f"Remito generado en {file_path}")

        # Abrir el archivo automáticamente
        try:
            os.startfile(file_path)
            self.carrito = []
            self.actualizar_carrito()

        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir el archivo: {str(e)}")

        
      

    def generar_presupuesto_excel(self):
        # Solicitar ubicación para guardar el archivo Excel
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            return

        # Cargar la plantilla de Excel para el presupuesto con openpyxl y obtener la hoja activa
        wb = load_workbook('./data/PLANTILLA PRESUPUESTO.xlsx')
        sheet = wb.active

        # Obtener la fecha actual
        fecha_actual = datetime.date.today().strftime("%d-%m-%Y")

        # Rellenar los datos generales del presupuesto
        # Fecha
        sheet.cell(row=3, column=6, value=fecha_actual)

        # Estilo de borde y relleno para las celdas aplicado con openpyxl
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
        blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

        # Rellenar la plantilla con los datos del presupuesto
        fila_inicial = 11  # Fila donde comienzan los productos en la plantilla
        subtotal = 0

        # Iterar sobre los productos en el carrito y hacer los cálculos necesarios
        for item in self.carrito:
            # Desempaquetar los datos de la lista del carrito en variables separadas 
            producto, cantidad, descuento, precio = item
            # Convertir la cantidad y el precio a entero y flotante
            cantidad = int(cantidad)
            precio = float(precio)
            # Calcular el precio total con descuento y sin IVA del producto
            precio_total = cantidad * precio * (1 - descuento / 100)
            precio_sin_iva = precio_total / 1.21
            # Convertir el descuento a porcentaje
            descuento_porcentaje = float(descuento) / 100

            # Escribir los datos en las celdas correspondientes
            sheet.cell(row=fila_inicial, column=2, value=cantidad)
            sheet.cell(row=fila_inicial, column=3, value=producto)
            sheet.cell(row=fila_inicial, column=4, value=precio)
            sheet.cell(row=fila_inicial, column=5, value=descuento_porcentaje).number_format = '0.00%'
            sheet.cell(row=fila_inicial, column=6, value=precio_total)

            # Aplicar borde a las celdas de la fila actual
            for col in range(2, 7):
                sheet.cell(row=fila_inicial, column=col).border = thin_border

            # Calcular el subtotal del presupuesto con los precios sin IVA
            subtotal += precio_sin_iva
            # Pasar a la siguiente fila
            fila_inicial += 1

        # Calcular los totales
        iva = subtotal * 0.21
        total = subtotal + iva

        # Escribir los totales en las celdas correspondientes
        sheet.cell(row=fila_inicial, column=5, value="Subtotal").font = Font(bold=True)
        sheet.cell(row=fila_inicial, column=6, value=subtotal)

        sheet.cell(row=fila_inicial + 1, column=5, value="IVA").font = Font(bold=True)
        sheet.cell(row=fila_inicial + 1, column=6, value=iva)

        sheet.cell(row=fila_inicial + 2, column=5, value="Total").font = Font(bold=True)
        sheet.cell(row=fila_inicial + 2, column=6, value=total)

        # Gracias por su confianza en la tercer columna
        sheet.cell(row=fila_inicial + 2, column=3, value="GRACIAS POR SU CONFIANZA").font = Font(bold=True)

        # Aplicar borde y relleno a las celdas de subtotal, iva y total
        for row in range(fila_inicial, fila_inicial + 3):
            for col in range(5, 7):
                cell = sheet.cell(row=row, column=col)
                cell.border = thin_border
                # Aplicar relleno azul a las celdas de subtotal, iva y total
                # cell.fill = blue_fill

        # Guardar el archivo Excel
        wb.save(file_path)
        messagebox.showinfo("Éxito", f"Presupuesto generado en {file_path}")

        # Abrir el archivo automáticamente
        try:
            os.startfile(file_path)
            self.carrito = []
            self.actualizar_carrito()

        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir el archivo: {str(e)}")

# Función principal para ejecutar la aplicación
# Si el script se ejecuta directamente, se crea una instancia de la clase PresupuestoApp y se llama al método mainloop
if __name__ == "__main__":
    app = PresupuestoApp()
    app.mainloop()
