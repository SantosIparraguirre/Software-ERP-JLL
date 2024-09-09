import datetime
import os
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

def generar_presupuesto_excel(cliente_var, carrito, session, Clientes, imprimir):
    # Obtener el cliente
    cliente = cliente_var.get()

    # Verificar si se seleccionó un cliente
    if not cliente:
        return

    # Obtener la ruta del escritorio del usuario actual
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")

    # Construir la ruta relativa a la carpeta "PRESUPUESTOS" en el escritorio
    nota_de_entrega_path = os.path.join(desktop_path, "PRESUPUESTOS")

    # Solicitar al usuario la ubicación donde guardar el presupuesto, con la carpeta por defecto PRESUPUESTOS
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        initialdir=nota_de_entrega_path,
        title="Guardar remito como..."
    )

    if not file_path:
        return

    # Cargar la plantilla de Excel para el presupuesto con openpyxl y obtener la hoja activa
    wb = load_workbook('./data/PLANTILLA PRESUPUESTO.xlsx')
    sheet = wb.active

    # Obtener la fecha actual
    fecha_actual = datetime.date.today().strftime("%d-%m-%Y")

    # Rellenar los datos generales del presupuesto
    # Fecha
    sheet.cell(row=3, column=6, value=fecha_actual)

    # Si el cliente es "Consumidor Final", no imputamos datos de cliente
    if cliente != "Consumidor Final":
        # Verificar si el cliente está en la base de datos
        cliente_db = session.query(Clientes).filter_by(nombre=cliente).first()
        if cliente_db:
            nombre_cliente = cliente_db.nombre
            cuit_cliente = cliente_db.cuit
            domicilio_cliente = cliente_db.direccion
            telefono_cliente = cliente_db.telefono

            # Escribir los datos del cliente en las celdas correspondientes
            sheet.cell(row=5, column=6, value=nombre_cliente)
            sheet.cell(row=6, column=6, value=cuit_cliente)
            sheet.cell(row=7, column=6, value=domicilio_cliente)
            sheet.cell(row=9, column=6, value=telefono_cliente)

    # Estilo de borde y relleno para las celdas aplicado con openpyxl
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
    blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    # Rellenar la plantilla con los datos del presupuesto
    fila_inicial = 11  # Fila donde comienzan los productos en la plantilla
    subtotal = 0

    # Iterar sobre los productos en el carrito y hacer los cálculos necesarios
    for item in carrito:
        # Desempaquetar los datos de la lista del carrito en variables separadas 
        producto, cantidad, descuento, precio = item
        # Convertir la cantidad y el precio a entero y flotante
        cantidad = float(cantidad)
        # Quitarle el signo $ al precio y la coma
        precio = precio[1:].replace(',', '') if precio else 0
        # Convertir el precio a float
        precio = float(precio) if precio else 0
        # Quitarle el signo % al descuento y convertirlo a float
        descuento = float(descuento[:-1]) if descuento else ''
        # Calcular el precio total con descuento y sin IVA del producto
        precio_total = cantidad * precio * (1 - descuento / 100 if descuento else 1)
        precio_sin_iva = precio_total / 1.21
        # Convertir el descuento a string y agregarle el signo de porcentaje
        descuento = f'{descuento}%' if descuento else ''
        
        # Escribir los datos en las celdas correspondientes
        sheet.cell(row=fila_inicial, column=2, value=cantidad)
        sheet.cell(row=fila_inicial, column=3, value=producto)
        sheet.cell(row=fila_inicial, column=4, value=precio)
        sheet.cell(row=fila_inicial, column=5, value=descuento)
        sheet.cell(row=fila_inicial, column=6, value=precio_total)

        # Aplicar borde a las celdas de la fila actual
        for col in range(2, 7):
            sheet.cell(row=fila_inicial, column=col).border = thin_border

        # Calcular el subtotal del presupuesto con los precios sin IVA
        subtotal += precio_sin_iva
        # Pasar a la siguiente fila
        fila_inicial += 1

    # Calcular los totales
    iva = subtotal * 0.21
    total = subtotal + iva

    # Escribir los totales en las celdas correspondientes
    sheet.cell(row=fila_inicial, column=5, value="Subtotal").font = Font(name='Arial', size= 12, bold=True)
    sheet.cell(row=fila_inicial, column=6, value=subtotal)

    sheet.cell(row=fila_inicial + 1, column=5, value="IVA").font = Font(name='Arial', size= 12, bold=True)
    sheet.cell(row=fila_inicial + 1, column=6, value=iva)

    sheet.cell(row=fila_inicial + 2, column=5, value="Total").font = Font(name='Arial', size= 12, bold=True)
    sheet.cell(row=fila_inicial + 2, column=6, value=total)

    # Gracias por su confianza en la tercer columna
    sheet.cell(row=fila_inicial + 2, column=3, value="GRACIAS POR SU CONFIANZA").font = Font(name='Arial', size=12, bold=True)

    # Aplicar borde y relleno a las celdas de subtotal, iva y total
    for row in range(fila_inicial, fila_inicial + 3):
        for col in range(5, 7):
            cell = sheet.cell(row=row, column=col)
            cell.border = thin_border

    # Copia para la empresa

    if fila_inicial <= 37:
        fila_inicial = 40
        # Fusionar celdas para el título
        sheet.merge_cells(start_row=fila_inicial, start_column=4, end_row=fila_inicial +1, end_column=6)
        # Titulo "PRESUPUESTO" con fuente 'Arial' de tamaño 27 y negrita
        sheet.cell(row=fila_inicial, column=4, value="PRESUPUESTO").font = Font(name='Arial', size=27, color="8db3e2")
        # Alinear a la derecha y arriba
        sheet.cell(row=fila_inicial, column=4).alignment = Alignment(horizontal='right', vertical='top')


        # Guardar el nombre del cliente en una variable
        nombre_cliente = cliente_var.get()

        # Si el cliente no es "Consumidor Final", rellenar los datos del cliente
        if nombre_cliente != "Consumidor Final":
            # Etiqueta "CLIENTE" y nombre del cliente alineado a la izquierda
            sheet.cell(row=fila_inicial, column=2, value=f"CLIENTE: {nombre_cliente}").alignment = Alignment(horizontal='left')

            # Etiqueta "Domicilio" y domicilio del cliente
            sheet.cell(row=fila_inicial + 1, column=2, value=f"DOMICILIO: {domicilio_cliente}").alignment = Alignment(horizontal='left')

            # Etiqueta "CUIT" y CUIT del cliente
            sheet.cell(row=fila_inicial + 2, column=2, value=f"CUIT: {cuit_cliente}").alignment = Alignment(horizontal='left')

            # Etiqueta "DNI" y espacio para completar
            sheet.cell(row=fila_inicial + 3, column=2, value=f"TELEFONO: {telefono_cliente}").alignment = Alignment(horizontal='left')

            # Etiqueta "Fecha"
            sheet.cell(row=fila_inicial + 3, column=5, value=f"FECHA: {fecha_actual}").alignment = Alignment(horizontal='left')

        else:
            # Si el cliente es "Consumidor Final", escribir "Consumidor Final" en la celda correspondiente
            sheet.cell(row=fila_inicial, column=2, value="CLIENTE: Consumidor Final").alignment = Alignment(horizontal='left')

            # Etiqueta "Fecha"
            sheet.cell(row=fila_inicial + 1, column=2, value=f"FECHA: {fecha_actual}").alignment = Alignment(horizontal='left')

            fila_inicial -= 2

        # Borde para las celdas
        borde = Border(left=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'),
                            right=Side(style='thin'))
        
        # CANTIDAD, DETALLE, PRECIO UD., DESCUENTO, IMPORTE C/IVA
        sheet.cell(row=fila_inicial + 5, column=2, value="CANTIDAD").border = borde
        sheet.cell(row=fila_inicial + 5, column=3, value="DETALLE").border = borde
        sheet.cell(row=fila_inicial + 5, column=4, value="PRECIO UD.").border = borde
        sheet.cell(row=fila_inicial + 5, column=5, value="DESCUENTO").border = borde
        sheet.cell(row=fila_inicial + 5, column=6, value="IMPORTE C/IVA").border = borde

        # Color
        color_fondo = PatternFill(start_color="dbe5f1", end_color="dbe5f1", fill_type="solid")

        # Aplicar color de fondo dbe5f1 a la fila de los encabezados
        for col in range(2, 7):
            sheet.cell(row=fila_inicial + 5, column=col).fill = color_fondo

        # Rellenar la plantilla con los datos del presupuesto
        fila_inicial += 6  # Fila donde comienzan los productos en la plantilla
        subtotal = 0

        # Iterar sobre los productos en el carrito y hacer los cálculos necesarios
        for item in carrito:
            # Desempaquetar los datos de la lista del carrito en variables separadas 
            producto, cantidad, descuento, precio = item
            # Convertir la cantidad y el precio a entero y flotante
            cantidad = float(cantidad)
            # Quitarle el signo $ al precio y la coma
            precio = precio[1:].replace(',', '') if precio else 0
            precio = float(precio) if precio else 0
            # Quitar el signo de porcentaje y convertir a flotante el descuento
            descuento = float(descuento[:-1]) if descuento else ''
            # Calcular el precio total con descuento y sin IVA del producto
            precio_total = cantidad * precio * (1 - descuento / 100 if descuento else 1)
            precio_sin_iva = precio_total / 1.21
            # Convertir el descuento a string y agregarle el signo de porcentaje
            descuento = f'{descuento}%' if descuento else ''
            
            # Escribir los datos en las celdas correspondientes
            sheet.cell(row=fila_inicial, column=2, value=cantidad)
            sheet.cell(row=fila_inicial, column=3, value=producto)
            sheet.cell(row=fila_inicial, column=4, value=precio)
            sheet.cell(row=fila_inicial, column=5, value=descuento)
            sheet.cell(row=fila_inicial, column=6, value=precio_total)

            # Aplicar borde a las celdas de la fila actual
            for col in range(2, 7):
                sheet.cell(row=fila_inicial, column=col).border = borde

            # Calcular el subtotal del presupuesto con los precios sin IVA
            subtotal += precio_sin_iva
            # Pasar a la siguiente fila
            fila_inicial += 1

        # Calcular los totales
        iva = subtotal * 0.21
        total = subtotal + iva

        # Escribir los totales en las celdas correspondientes
        sheet.cell(row=fila_inicial, column=5, value="Subtotal").font = Font(name='Arial', size= 12, bold=True)
        sheet.cell(row=fila_inicial, column=6, value=subtotal)

        sheet.cell(row=fila_inicial + 1, column=5, value="IVA").font = Font(name='Arial', size= 12, bold=True)
        sheet.cell(row=fila_inicial + 1, column=6, value=iva)

        sheet.cell(row=fila_inicial + 2, column=5, value="Total").font = Font(name='Arial', size= 12, bold=True)
        sheet.cell(row=fila_inicial + 2, column=6, value=total)

        # Gracias por su confianza en la tercer columna
        sheet.cell(row=fila_inicial + 2, column=3, value="GRACIAS POR SU CONFIANZA").font = Font(name='Arial', size=12, bold=True)

        # Aplicar borde y relleno a las celdas de subtotal, iva y total
        for row in range(fila_inicial, fila_inicial + 3):
            for col in range(5, 7):
                cell = sheet.cell(row=row, column=col)
                cell.border = borde

        # Guardar el archivo Excel
        wb.save(file_path)
        messagebox.showinfo("Éxito", f"Presupuesto generado en {file_path}")

        # Si clickeó en "Imprimir Presupuesto"
        if imprimir:
            try:
                os.startfile(file_path, "print")
            except Exception as e:
                messagebox.showerror("Error", f"No se pudo imprimir el presupuesto: {str(e)}")

        else:
            # Abrir el archivo automáticamente
            try:
                os.startfile(file_path)

            except Exception as e:
                messagebox.showerror("Error", f"No se pudo abrir el archivo: {str(e)}")