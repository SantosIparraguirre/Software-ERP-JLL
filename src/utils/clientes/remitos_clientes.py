import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from database import Clientes, Remitos, DetallesRemitos, session  
from datetime import datetime
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

ventana_remitos = None

def ver_remitos(self):
    # Obtener el cliente seleccionado en la tabla
    seleccion = self.clientes_tree.selection()
    if not seleccion:
        messagebox.showerror("Error", "Selecciona un cliente.")
        return
    
    nombre = self.clientes_tree.item(seleccion)['values'][0]
    abrir_ventana_remitos(self, nombre)

def abrir_ventana_remitos(self, nombre):
    global ventana_remitos

    # Si la ventana de remitos ya está abierta, destruir la ventana actual y crear una nueva
    if ventana_remitos and tk.Toplevel.winfo_exists(ventana_remitos):
        ventana_remitos.destroy()

    # Crear una ventana secundaria para ver los remitos del cliente
    ventana_remitos = tk.Toplevel()
    ventana_remitos.title("Remitos de " + nombre)
    ventana_remitos.geometry("900x600")

    # Crear un marco para los remitos
    frame_remitos = tk.Frame(ventana_remitos)
    frame_remitos.pack(fill='both', expand=True)

    # Crear scrollbar para el Treeview de remitos
    scrollbar_remitos = ttk.Scrollbar(frame_remitos, orient="vertical")

    # Crear una tabla para mostrar los remitos del cliente
    remitos_tree = ttk.Treeview(frame_remitos, columns=('ID', 'Fecha', 'Fecha Pago', 'Total', 'Pago', 'Observacion'), show='headings', yscrollcommand=scrollbar_remitos.set)
    remitos_tree.heading('ID', text='ID')
    remitos_tree.heading('Fecha', text='Fecha')
    remitos_tree.heading('Fecha Pago', text='Fecha Pago')
    remitos_tree.heading('Total', text='Total')
    remitos_tree.heading('Pago', text='Pago')
    remitos_tree.heading('Observacion', text='Observacion')
    remitos_tree.column('ID', width=30, anchor='center')
    remitos_tree.column('Fecha', width=150, anchor='center')
    remitos_tree.column('Fecha Pago', width=150, anchor='center')
    remitos_tree.column('Total', width=100, anchor='center')
    remitos_tree.column('Pago', width=100, anchor='center')
    remitos_tree.column('Observacion', width=200, anchor='center')

    # Vincular el scrollbar al Treeview de remitos
    scrollbar_remitos.config(command=remitos_tree.yview)
    
    # Posicionar el treeview y las scrollbars
    remitos_tree.grid(row=0, column=0, sticky='nsew')
    scrollbar_remitos.grid(row=0, column=1, sticky='ns')

    # Configurar el frame para que se expanda
    frame_remitos.grid_rowconfigure(0, weight=1)
    frame_remitos.grid_columnconfigure(0, weight=1)

    # Obtener el cliente de la base de datos por el nombre
    cliente = session.query(Clientes).filter_by(nombre=nombre).first()
    for remito in cliente.remitos:
        # Formatear la fecha sin los decimales de los segundos
        fecha_formateada = remito.fecha.strftime("%d/%m/%Y %H:%M:%S") if remito.fecha else ""
        fecha_pago_formateada = remito.fecha_pago.strftime("%d/%m/%Y %H:%M:%S") if remito.fecha_pago else ""
        # Insertar en el treeview los valores del remito, formateando el total y el pago como moneda
        if remito.pago != "NO" and remito.pago != "SI":
            pago_formateado = f"${float(remito.pago):,.2f}"
        else:
            pago_formateado = remito.pago
        remitos_tree.insert('', 'end', values=(remito.id, fecha_formateada, fecha_pago_formateada, f"${remito.total:,.2f}", pago_formateado, remito.observacion))
    
    # Frame para los botones
    frame_botones = tk.Frame(ventana_remitos)
    frame_botones.pack(pady=10)

    # Botón para unir remitos
    unir_remitos_button = ttk.Button(frame_botones, text="Remito General", command=lambda: unir_remitos(remitos_tree, detalles_tree, ventana_remitos))
    unir_remitos_button.grid(row=0, column=0, padx=5)

    # Botón para eliminar remitos
    eliminar_remito_button = ttk.Button(frame_botones, text="Eliminar Remito", command=lambda: eliminar_remito(remitos_tree, ventana_remitos))
    eliminar_remito_button.grid(row=0, column=1, padx=5)

    # Botón para modificar un detalle del remito
    # El botón toma como argumento la fila seleccionada en el Treeview de detalles
    modificar_detalle_button = ttk.Button(frame_botones, text="Modificar Detalle", command=lambda: modificar_detalle(remitos_tree, detalles_tree, ventana_remitos))
    modificar_detalle_button.grid(row=0, column=2, padx=5)

    # Botón para agregar un detalle al remito
    # El botón toma como argumento el ID del remito seleccionado
    agregar_detalle_button = ttk.Button(frame_botones, text="Agregar Detalle", command=lambda: agregar_detalle(remitos_tree, ventana_remitos, detalles_tree))
    agregar_detalle_button.grid(row=0, column=3, padx=5)

    # Botón para eliminar un detalle del remito
    # El botón toma como argumento la fila seleccionada en el Treeview de detalles
    eliminar_detalle_button = ttk.Button(frame_botones, text="Eliminar Detalle", command=lambda: eliminar_detalle(remitos_tree, detalles_tree, ventana_remitos))
    eliminar_detalle_button.grid(row=0, column=4, padx=5)

    # Crear un marco para los detalles del remito
    frame_detalles = tk.Frame(ventana_remitos)
    frame_detalles.pack(fill='both', expand=True)

    # Crear scrollbar para el Treeview de detalles
    scrollbar_detalles = ttk.Scrollbar(frame_detalles, orient="vertical")

    # Crear un Treeview vacío para los detalles del remito seleccionado
    detalles_tree = ttk.Treeview(frame_detalles, columns=('ID', 'Producto', 'Cantidad', 'Precio Unitario', 'Descuento', 'Total'), show='headings', yscrollcommand=scrollbar_detalles.set)
    detalles_tree.heading('ID', text='ID')
    detalles_tree.heading('Producto', text='Producto')
    detalles_tree.heading('Cantidad', text='Cantidad')
    detalles_tree.heading('Precio Unitario', text='Precio Unitario')
    detalles_tree.heading('Descuento', text='Descuento')
    detalles_tree.heading('Total', text='Total')
    detalles_tree.column('ID', width=0, stretch=tk.NO)
    detalles_tree.column('Producto', width=250, anchor='center')
    detalles_tree.column('Cantidad', width=50, anchor='center')
    detalles_tree.column('Precio Unitario', width=150, anchor='center')
    detalles_tree.column('Descuento', width=100, anchor='center')
    detalles_tree.column('Total', width=150, anchor='center')

    # Vincular el scrollbar al Treeview de detalles
    scrollbar_detalles.config(command=detalles_tree.yview)

    # Posicionar el treeview y el scrollbar
    detalles_tree.grid(row=0, column=0, sticky='nsew')
    scrollbar_detalles.grid(row=0, column=1, sticky='ns')

    # Configurar el frame para que se expanda
    frame_detalles.grid_rowconfigure(0, weight=1)
    frame_detalles.grid_columnconfigure(0, weight=1)

    # Vincular el evento de selección del Treeview de remitos a la función ver_detalles_remito
    remitos_tree.bind('<<TreeviewSelect>>', lambda event: ver_detalles_remito(remitos_tree, detalles_tree))

def unir_remitos(remitos_tree, detalles_tree, ventana_remitos):
    # Obtener los IDs de los remitos seleccionados
    selected_items = remitos_tree.selection()

    # Validar que se hayan seleccionado dos o más remitos
    if len(selected_items) < 2 or not selected_items:
        messagebox.showerror("Error", "Por favor, seleccione dos o más remitos para unir.", parent=ventana_remitos)
        return
    
    # Mostrar un mensaje de confirmación
    confirmacion = messagebox.askyesno("Confirmación", "¿Estás seguro de unir los remitos seleccionados?", parent=ventana_remitos)
    if not confirmacion:
        return

    # Crear una lista para almacenar los detalles de los remitos seleccionados
    remitos_unidos = []
    total_general = 0

    for item in selected_items:
        remito_id = remitos_tree.item(item, 'values')[0]
        remito = session.query(Remitos).filter_by(id=remito_id).first()
        if remito:
            # Recorrer los detalles del remito y combinarlos
            for detalle in remito.detalles:
                # Buscar si el producto ya está en la lista de detalles unidos
                encontrado = False
                for detalle_unido in remitos_unidos:
                    if detalle.producto == detalle_unido["producto"]:
                        # # Calcular el nuevo precio unitario ponderado
                        # detalle_unido["precio_unitario"] = (
                        #     (detalle_unido["precio_unitario"] * detalle_unido["cantidad"] + detalle.precio_unitario * detalle.cantidad)
                        #     / (detalle_unido["cantidad"] + detalle.cantidad)
                        # )

                        # # Calcular el nuevo descuento ponderado
                        # detalle_unido["descuento"] = (
                        #     (detalle_unido["descuento"] * detalle_unido["cantidad"] + detalle.descuento * detalle.cantidad)
                        #     / (detalle_unido["cantidad"] + detalle.cantidad)
                        # )
                        # Si ya está, sumar la cantidad y el total
                        detalle_unido["cantidad"] += detalle.cantidad
                        detalle_unido["total"] += detalle.total
                        encontrado = True
                        break
                
                if not encontrado:
                    # Si no está, hacer una copia del detalle y agregarlo a la lista
                    detalle_copia = {
                        "producto": detalle.producto,
                        "cantidad": detalle.cantidad,
                        # "precio_unitario": detalle.precio_unitario,
                        # "descuento": detalle.descuento,
                        "total": detalle.total
                    }
                    remitos_unidos.append(detalle_copia)

            # Sumar el total del remito al total general
            total_general += remito.total

    # Crear un nuevo remito con los detalles combinados
    remito_general = Remitos(
        id_cliente=remito.id_cliente,
        fecha=datetime.now(),
        fecha_pago=None,
        total=total_general,
        pago="SI",
        observacion="Remito general",
    )
    session.add(remito_general)
    session.commit()

    # Agregar los detalles al nuevo remito general
    for detalle_unido in remitos_unidos:
        nuevo_detalle = DetallesRemitos(
            id_remito=remito_general.id,
            producto=detalle_unido["producto"],
            cantidad=detalle_unido["cantidad"],
            # precio_unitario=detalle_unido["precio_unitario"],
            # descuento=detalle_unido["descuento"],
            total=detalle_unido["total"],
        )
        session.add(nuevo_detalle)

    session.commit()

    messagebox.showinfo("Éxito", f"Remito general creado con un total de ${total_general:,.2f}", parent=ventana_remitos)

    # Llamar a la función generar_remito con el remito general
    exportar_remito_excel(remito_general)

    # Actualizar los remitos
    actualizar_remitos(remitos_tree, remito.cliente.nombre)

def exportar_remito_excel(remito):
    # Obtener la ruta del escritorio del usuario actual
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")

    # Construir la ruta relativa a la carpeta "REMITOS" en el escritorio
    nota_de_entrega_path = os.path.join(desktop_path, "REMITOS")

    # Solicitar al usuario la ubicación donde guardar el remito, con la carpeta por defecto NOTA DE ENTREGA
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        initialdir=nota_de_entrega_path,
        title="Guardar remito como..."
    )

    if not file_path:
        return
    
    # Cargar el archivo Excel
    wb = load_workbook("./data/PLANTILLA REMITO GENERAL.xlsx")
    sheet = wb.active

    # Obtener los datos del cliente
    cliente = remito.cliente

    nombre = cliente.nombre if cliente.nombre else ""
    direccion = cliente.direccion if cliente.direccion else ""
    cuit = cliente.cuit if cliente.cuit else ""
    telefono = cliente.telefono if cliente.telefono else ""

    # Definir estilos
    bold_italic_font = Font(name='Arial', size=12, bold=True, italic=True, color="8EB4E3")
    bold_font = Font(name='Arial', size=12, bold=True)
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    # Obtener la fecha actual (dia, mes y año)
    fecha = datetime.now().strftime("%d/%m/%Y")

    # Escribir los datos del cliente y la fecha en el remito
    sheet["E3"] = fecha
    sheet["E5"] = nombre
    sheet["E6"] = cuit
    sheet["E7"] = direccion
    sheet["E9"] = telefono

    # Obtener los detalles del remito
    detalles = remito.detalles

    # Escribir los detalles en el remito
    row = 11
    for detalle in detalles:
        sheet[f"B{row}"] = detalle.cantidad
        sheet[f"B{row}"].alignment = Alignment(horizontal='center')

        sheet[f"C{row}"] = detalle.producto

        sheet[f"E{row}"] = detalle.total
        
        row += 1

    # Calcular subtotal, IVA y total
    subtotal = sum(detalle.total for detalle in detalles) / 1.21
    iva = subtotal * 0.21
    total = subtotal + iva

    # Escribir los montos en el remito con negrita y cursiva + color
    sheet[f"D{row + 1}"] = "SUBTOTAL"
    sheet[f"D{row + 1}"].font = bold_italic_font

    sheet[f"E{row + 1}"] = subtotal

    sheet[f"D{row + 2}"] = "IVA"
    sheet[f"D{row + 2}"].font = bold_italic_font

    sheet[f"E{row + 2}"] = iva

    sheet[f"D{row + 3}"] = "TOTAL"
    sheet[f"D{row + 3}"].font = bold_italic_font

    sheet[f"E{row + 3}"] = total
    sheet[f"E{row + 3}"].font = bold_font

    # "Gracias por su confianza"
    sheet[f"B{row + 3}"] = "GRACIAS POR SU CONFIANZA"
    sheet[f"B{row + 3}"].font = bold_italic_font

    # Copia para la empresa
    if row + 5 <= 43:
        # Duplicar la información en la misma hoja, a partir de la fila 44
        row = 44

        # "REMITO" con tamaño 27 y color 8eb4e3 en columna E41
        sheet[f"E{row}"] = "REMITO"
        sheet[f"E{row}"].font = Font(name='Arial', size=27, color="8EB4E3")

        # "Documento no valido como factura" con tamaño 8 en columna E42
        sheet[f"E{row + 1}"] = "Documento no válido como factura"
        sheet[f"E{row + 1}"].font = Font(name='Arial', size=8)
        # Alinear arriba
        sheet[f"E{row + 1}"].alignment = Alignment(vertical='top', horizontal='center')

        # "Fecha" en columna B43
        sheet[f"B{row}"] = f"Fecha: {fecha}"

        # "Cliente" en columna B44
        sheet[f"B{row + 2}"] = f"Cliente: {nombre}"

        # "CUIT" en columna B45
        sheet[f"B{row + 3}"] = f"CUIT: {cuit}"

        # "Dirección" en columna D44
        sheet[f"D{row + 2}"] = f"Dirección: {direccion}"

        # "Teléfono" en columna D45
        sheet[f"D{row + 3}"] = f"Teléfono: {telefono}"

        # "CANTIDAD" en columna B46 con relleno dce6f2
        sheet[f"B{row + 4}"] = "CANTIDAD"
        sheet[f"B{row + 4}"].fill = PatternFill(start_color="dce6f2", end_color="dce6f2", fill_type="solid")
        # Centrar
        sheet[f"B{row + 4}"].alignment = Alignment(horizontal='center')

        # Rellenar la celda vacía en C46 con relleno dce6f2
        sheet[f"C{row + 4}"].fill = PatternFill(start_color="dce6f2", end_color="dce6f2", fill_type="solid")
        
        # "DESCRIPCIÓN" en columna D46 con relleno dce6f2
        sheet[f"D{row + 4}"] = "DESCRIPCIÓN"
        sheet[f"D{row + 4}"].fill = PatternFill(start_color="dce6f2", end_color="dce6f2", fill_type="solid")

        # "IMPORTE C/ IVA" en columna E46 con relleno dce6f2
        sheet[f"E{row + 4}"] = "IMPORTE C/ IVA"
        sheet[f"E{row + 4}"].fill = PatternFill(start_color="dce6f2", end_color="dce6f2", fill_type="solid")

        # Escribir los detalles en la copia del remito
        row += 5
        for detalle in detalles:
            sheet[f"B{row}"] = detalle.cantidad
            sheet[f"B{row}"].alignment = Alignment(horizontal='center')

            sheet[f"C{row}"] = detalle.producto

            sheet[f"E{row}"] = detalle.total

            row += 1
        
        # Calcular subtotal, IVA y total
        subtotal = sum(detalle.total for detalle in detalles) / 1.21
        iva = subtotal * 0.21
        total = subtotal + iva

        # Escribir los montos en la copia del remito
        sheet[f"D{row + 1}"] = "SUBTOTAL"
        sheet[f"D{row + 1}"].font = bold_italic_font
        sheet[f"D{row + 1}"].alignment = Alignment(horizontal='right')

        sheet[f"E{row + 1}"] = subtotal

        sheet[f"D{row + 2}"] = "IVA"
        sheet[f"D{row + 2}"].font = bold_italic_font
        sheet[f"D{row + 2}"].alignment = Alignment(horizontal='right')

        sheet[f"E{row + 2}"] = iva

        sheet[f"D{row + 3}"] = "TOTAL"
        sheet[f"D{row + 3}"].font = bold_italic_font
        sheet[f"D{row + 3}"].alignment = Alignment(horizontal='right')

        sheet[f"E{row + 3}"] = total
        sheet[f"E{row + 3}"].font = bold_font

    # Guardar el archivo Excel
    wb.save(file_path)

    # Mostrar un mensaje de éxito
    messagebox.showinfo("Éxito", f"Remito guardado exitosamente en {file_path}")

    # Abrir el archivo Excel
    os.startfile(file_path)

def ver_detalles_remito(remitos_tree, detalles_tree):
    # Obtener el remito seleccionado en la tabla
    seleccion = remitos_tree.selection()
    if not seleccion or len(seleccion) > 1:
        return

    # Obtener el ID del remito seleccionado
    ID = remitos_tree.item(seleccion)['values'][0]
    # Llamar a la función mostrar_detalles_remito con el ID del remito y el Treeview de detalles
    mostrar_detalles_remito(ID, detalles_tree)

def mostrar_detalles_remito(ID, detalles_tree):
    # Limpiar el Treeview de detalles previo
    for item in detalles_tree.get_children():
        detalles_tree.delete(item)

    # Obtener el remito de la base de datos por la ID
    remito = session.query(Remitos).filter_by(id=ID).first()
    for detalle in remito.detalles:
        # Si es un remito general, no mostrar el precio unitario ni el descuento
        if remito.observacion == 'Remito general':
            detalles_tree.insert('', 'end', values=(detalle.id, detalle.producto, detalle.cantidad, '', '', f"${detalle.total:,.2f}"))
        # Si no, mostrar todos los detalles
        else:
            detalles_tree.insert('', 'end', values=(detalle.id, detalle.producto, detalle.cantidad, f"${detalle.precio_unitario:,.2f}", detalle.descuento, f"${detalle.total:,.2f}"))
        

def eliminar_remito(remitos_tree, ventana_remitos):
    # Obtener el remito seleccionado en la tabla
    seleccion = remitos_tree.selection()
    if not seleccion:
        messagebox.showerror("Error", "Selecciona un remito.", parent=ventana_remitos)
        return

    # Obtener el ID del remito seleccionado
    ID = remitos_tree.item(seleccion)['values'][0]
    # Buscar el remito en la base de datos por la ID
    remito = session.query(Remitos).filter_by(id=ID).first()
    # Preguntar al usuario si está seguro de eliminar el remito
    confirmacion = messagebox.askyesno("Confirmación", f"¿Estás seguro de eliminar el remito {ID}?", parent=ventana_remitos)
    if not confirmacion:
        return
    
    # Obtener el nombre del cliente del remito eliminado
    nombre_cliente = remito.cliente.nombre
    
    # Eliminar los detalles del remito
    session.query(DetallesRemitos).filter_by(id_remito=ID).delete()

    # Eliminar el remito
    session.query(Remitos).filter_by(id=ID).delete()

    # Confirmar la transacción
    session.commit()

    actualizar_remitos(remitos_tree, nombre_cliente)

    # Mostrar un mensaje de éxito
    messagebox.showinfo("Éxito", "Remito eliminado exitosamente.", parent=ventana_remitos)

def actualizar_remitos(remitos_tree, nombre):
    # Limpiar el Treeview de remitos previo
    for item in remitos_tree.get_children():
        remitos_tree.delete(item)

    # Obtener el cliente de la base de datos por el nombre
    cliente = session.query(Clientes).filter_by(nombre=nombre).first()

    # Insertar en el Treeview los valores actualizados del remito
    for remito in cliente.remitos:
        fecha_formateada = remito.fecha.strftime("%d/%m/%Y %H:%M:%S") if remito.fecha else ""
        fecha_pago_formateada = remito.fecha_pago.strftime("%d/%m/%Y %H:%M:%S") if remito.fecha_pago else ""
        if remito.pago != "NO" and remito.pago != "SI":
            pago_formateado = f"${float(remito.pago):,.2f}"
        else:
            pago_formateado = remito.pago
        remitos_tree.insert('', 'end', values=(remito.id, fecha_formateada, fecha_pago_formateada, f"${remito.total:,.2f}", pago_formateado, remito.observacion))

def modificar_detalle(remitos_tree, detalles_tree, ventana_remitos):
    # Obtener el remito seleccionado en la tabla
    seleccion_remito = remitos_tree.selection()
    if not seleccion_remito:
        messagebox.showerror("Error", "Selecciona un remito.", parent=ventana_remitos)
        return

    # Obtener el ID del remito seleccionado
    ID = remitos_tree.item(seleccion_remito)['values'][0]
    # Buscar el remito en la base de datos por la ID
    remito = session.query(Remitos).filter_by(id=ID).first()

    # Obtener el detalle seleccionado en la tabla
    seleccion_detalle = detalles_tree.selection()
    if not seleccion_detalle:
        messagebox.showerror("Error", "Selecciona un detalle.", parent=ventana_remitos)
        return

    # Obtener el ID del detalle seleccionado directo desde la base de datos
    ID_detalle = detalles_tree.item(seleccion_detalle)['values'][0]
    # Buscar el detalle en la base de datos por la ID
    detalle = session.query(DetallesRemitos).filter_by(id=ID_detalle).first()

    # Crear una ventana secundaria para modificar el detalle del remito
    ventana_mod_detalle = tk.Toplevel()
    ventana_mod_detalle.title("Modificar Detalle")
    ventana_mod_detalle.geometry("300x200")

    # Crear etiquetas y campos de entrada para los datos del detalle
    ttk.Label(ventana_mod_detalle, text="Producto:").grid(row=0, column=0, padx=5, pady=5)
    producto_var = tk.StringVar(value=detalle.producto)
    ttk.Entry(ventana_mod_detalle, textvariable=producto_var).grid(row=0, column=1, padx=5, pady=5)

    ttk.Label(ventana_mod_detalle, text="Cantidad:").grid(row=1, column=0, padx=5, pady=5)
    cantidad_var = tk.StringVar(value=detalle.cantidad)
    ttk.Entry(ventana_mod_detalle, textvariable=cantidad_var).grid(row=1, column=1, padx=5, pady=5)

    ttk.Label(ventana_mod_detalle, text="Precio Unitario:").grid(row=2, column=0, padx=5, pady=5)
    precio_var = tk.StringVar(value=detalle.precio_unitario)
    ttk.Entry(ventana_mod_detalle, textvariable=precio_var).grid(row=2, column=1, padx=5, pady=5)

    ttk.Label(ventana_mod_detalle, text="Descuento:").grid(row=3, column=0, padx=5, pady=5)
    descuento_var = tk.StringVar(value=detalle.descuento)
    ttk.Entry(ventana_mod_detalle, textvariable=descuento_var).grid(row=3, column=1, padx=5, pady=5)

    # Botón para guardar los cambios

    ttk.Button(ventana_mod_detalle, text="Guardar Cambios", 
               command=lambda: guardar_cambios_detalle(ID, ID_detalle, producto_var.get(), cantidad_var.get(), precio_var.get(), descuento_var.get(), remitos_tree, ventana_mod_detalle, detalles_tree)).grid(row=4, column=1, pady=10)
    
def guardar_cambios_detalle(ID_remito, ID_detalle, producto, cantidad, precio, descuento, remitos_tree, ventana_mod_detalle, detalles_tree):
    # Buscar el remito en la base de datos por la ID
    remito = session.query(Remitos).filter_by(id=ID_remito).first()
    # Buscar el detalle en la base de datos por la ID
    detalle = session.query(DetallesRemitos).filter_by(id=ID_detalle).first()

    # Actualizar los datos del detalle con los nuevos valores
    detalle.producto = producto
    detalle.cantidad = cantidad
    detalle.precio_unitario = precio
    detalle.descuento = descuento
    detalle.total = float(cantidad) * float(precio) * (1 - float(descuento) / 100)

    # Actualizar el total del remito
    remito.total = sum([detalle.total for detalle in remito.detalles])

    # Confirmar la transacción
    session.commit()

    # Mostrar un mensaje de éxito
    messagebox.showinfo("Éxito", "Detalle modificado exitosamente.", parent=ventana_mod_detalle)

    # Cerrar la ventana secundaria
    ventana_mod_detalle.destroy()

    # Actualizar los remitos
    actualizar_remitos(remitos_tree, remito.cliente.nombre)

    # Actualizar los detalles del remito
    mostrar_detalles_remito(ID_remito, detalles_tree)

def agregar_detalle(remitos_tree, ventana_remitos, detalles_tree):
    # Obtener el remito seleccionado en la tabla
    seleccion = remitos_tree.selection()
    if not seleccion:
        messagebox.showerror("Error", "Selecciona un remito.", parent=ventana_remitos)
        return

    # Obtener el ID del remito seleccionado
    ID = remitos_tree.item(seleccion)['values'][0]

    # Crear una ventana secundaria para agregar un detalle al remito
    ventana_agregar_detalle = tk.Toplevel()
    ventana_agregar_detalle.title("Agregar Detalle")
    ventana_agregar_detalle.geometry("300x200")

    # Crear etiquetas y campos de entrada para los datos del detalle
    ttk.Label(ventana_agregar_detalle, text="Producto:").grid(row=0, column=0, padx=5, pady=5)
    producto_var = tk.StringVar()
    ttk.Entry(ventana_agregar_detalle, textvariable=producto_var).grid(row=0, column=1, padx=5, pady=5)

    ttk.Label(ventana_agregar_detalle, text="Cantidad:").grid(row=1, column=0, padx=5, pady=5)
    cantidad_var = tk.StringVar()
    ttk.Entry(ventana_agregar_detalle, textvariable=cantidad_var).grid(row=1, column=1, padx=5, pady=5)

    ttk.Label(ventana_agregar_detalle, text="Precio Unitario:").grid(row=2, column=0, padx=5, pady=5)
    precio_var = tk.StringVar()
    ttk.Entry(ventana_agregar_detalle, textvariable=precio_var).grid(row=2, column=1, padx=5, pady=5)

    ttk.Label(ventana_agregar_detalle, text="Descuento:").grid(row=3, column=0, padx=5, pady=5)
    descuento_var = tk.StringVar()
    ttk.Entry(ventana_agregar_detalle, textvariable=descuento_var).grid(row=3, column=1, padx=5, pady=5)

    # Botón para guardar el detalle
    ttk.Button(ventana_agregar_detalle, text="Guardar Detalle", 
               command=lambda:
                guardar_detalle(ID, producto_var.get(), cantidad_var.get(), precio_var.get(), descuento_var.get(), ventana_remitos, ventana_agregar_detalle, detalles_tree, remitos_tree)).grid(row=4, column=1, pady=10)
    
def guardar_detalle(ID, producto, cantidad, precio, descuento, ventana_remitos, ventana_agregar_detalle, detalles_tree, remitos_tree):
    # Buscar el remito en la base de datos por la ID
    remito = session.query(Remitos).filter_by(id=ID).first()

    # Crear un nuevo detalle con los datos ingresados
    detalle = DetallesRemitos(
        id_remito=ID,
        producto=producto,
        cantidad=cantidad,
        precio_unitario=precio,
        descuento=descuento,
        total=int(cantidad) * float(precio) * (1 - float(descuento) / 100 if descuento else 0)
    )

    # Agregar el detalle al remito
    remito.detalles.append(detalle)

    # Actualizar el total del remito
    remito.total = sum([detalle.total for detalle in remito.detalles])

    # Confirmar la transacción
    session.commit()

    # Mostrar un mensaje de éxito
    messagebox.showinfo("Éxito", "Detalle agregado exitosamente.", parent=ventana_remitos)

    # Cerrar la ventana secundaria
    ventana_agregar_detalle.destroy()

    # Actualizar los detalles del remito
    mostrar_detalles_remito(ID, detalles_tree)

    # Actualizar los remitos
    actualizar_remitos(remitos_tree, remito.cliente.nombre)

def eliminar_detalle(remitos_tree, detalles_tree, ventana_remitos):
    # Obtener el detalle seleccionado en la tabla
    seleccion_detalle = detalles_tree.selection()

    if not seleccion_detalle:
        messagebox.showerror("Error", "Selecciona un detalle.", parent=ventana_remitos)
        return
    
    # Obtener el ID del detalle seleccionado
    ID_detalle = detalles_tree.item(seleccion_detalle)['values'][0]
    # Buscar el detalle en la base de datos por la ID
    detalle = session.query(DetallesRemitos).filter_by(id=ID_detalle).first()
    # Buscar el remito en la base de datos por la ID del detalle
    remito = session.query(Remitos).filter_by(id=detalle.id_remito).first()

    # Preguntar al usuario si está seguro de eliminar el detalle
    confirmacion = messagebox.askyesno("Confirmación", "¿Estás seguro de eliminar el detalle?", parent=ventana_remitos)
    if not confirmacion:
        return
    
    # Eliminar el detalle
    session.delete(detalle)

    # Confirmar la transacción
    session.commit()

    # Actualizar el total del remito
    remito.total = sum([detalle.total for detalle in remito.detalles])

    # Confirmar la transacción
    session.commit()

    # Mostrar un mensaje de éxito
    messagebox.showinfo("Éxito", "Detalle eliminado exitosamente.", parent=ventana_remitos)

    # Actualizar los detalles del remito
    mostrar_detalles_remito(remito.id, detalles_tree)

    # Actualizar los remitos
    actualizar_remitos(remitos_tree, remito.cliente.nombre)

